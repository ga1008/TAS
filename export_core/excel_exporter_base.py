# excel_exporter_base.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BaseExcelExporter:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

        # === 字体常量 ===
        self.FONT_NAME_SONG = '宋体'
        self.FONT_NAME_HEI = '黑体'

        # === 常用边框样式 ===
        self.BORDER_THIN = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def setup_page_layout(self, top=1.4, bottom=1.9, left=1.3, right=1.3, header=0.8, footer=0.8):
        """通用页面设置 (单位：英寸，openpyxl 默认使用英寸，需注意转换，这里按题目要求数值设置)"""
        # openpyxl 的 page_margins 单位通常是英寸。
        # 如果题目要求的 1.4 是厘米，需要 / 2.54。
        # 假设题目要求的单位是厘米(cm)，这是中文文档的惯例。
        cm_to_inch = 1 / 2.54
        self.ws.page_margins.top = top * cm_to_inch
        self.ws.page_margins.bottom = bottom * cm_to_inch
        self.ws.page_margins.left = left * cm_to_inch
        self.ws.page_margins.right = right * cm_to_inch
        self.ws.page_margins.header = header * cm_to_inch
        self.ws.page_margins.footer = footer * cm_to_inch

        # 设置纸张大小为 A4
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.page_setup.paperSize = self.ws.page_setup.PAPERSIZE_A4

    def set_cell_style(self, cell, font_name='宋体', font_size=11, bold=False, align_h='center', align_v='center',
                       border=True):
        """统一单元格样式设置"""
        cell.font = Font(name=font_name, size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=True)
        if border:
            cell.border = self.BORDER_THIN

    def set_row_height(self, row_idx, height):
        """设置行高 (单位：磅 points)"""
        self.ws.row_dimensions[row_idx].height = height

    def set_col_width(self, col_idx, width):
        """设置列宽 (单位：字符数)"""
        col_letter = get_column_letter(col_idx)
        self.ws.column_dimensions[col_letter].width = width
