import json

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from export_core.base_template import BaseExportTemplate
from export_core.excel_exporter_base import BaseExcelExporter
from extensions import db  # 引入数据库以查询成绩


class MachineTestScoreExporter(BaseExportTemplate, BaseExcelExporter):
    ID = "score_sheet_machinetest"
    NAME = "机试考核登分表 (Excel)"
    DESCRIPTION = "广西外国语学院标准机试/作品设计考核登分表，自动拉取班级成绩。"
    FILE_EXTENSION = "xlsx"

    # === UI Schema ===
    UI_SCHEMA = [
        {"name": "course_name", "label": "课程名称", "type": "text", "auto_fill_key": "course_name",
         "placeholder": "如：Python程序设计"},
        {"name": "course_code", "label": "课程编号", "type": "text", "auto_fill_key": "course_code",
         "placeholder": "如：E020001B4"},
        {"name": "class_name", "label": "专业年级班级", "type": "text", "auto_fill_key": "class_name",
         "placeholder": "必须与系统内的班级名一致，用于匹配成绩"},
        {"name": "teacher_name", "label": "授课教师", "type": "text", "auto_fill_key": "teacher"},

        # 题目配置：使用 JSON 格式或简易文本格式让用户输入题目分值
        {"name": "questions_config", "label": "题目分值配置", "type": "text",
         "placeholder": "格式：题号:分值，用逗号分隔。例：一:20,二:30,三:20,四:30",
         "value": "一:20,二:30,三:20,四:30"},

        {"name": "total_score", "label": "卷面总分", "type": "number", "value": 100}
    ]

    def __init__(self):
        BaseExcelExporter.__init__(self)

    def generate(self, content, meta_info, form_data, output_path):
        # 1. 解析参数
        course_name = form_data.get('course_name', '')
        course_code = form_data.get('course_code', '')
        class_name = form_data.get('class_name', '')
        teacher_name = form_data.get('teacher_name', '')
        total_score = form_data.get('total_score', '100')

        # 解析题目配置
        q_config_str = form_data.get('questions_config', "一:20,二:30,三:20,四:30")
        # 格式化为 list: [{'name': '一', 'score': '20'}, ...]
        questions = []
        try:
            for item in q_config_str.split(','):
                parts = item.split(':')
                if len(parts) >= 2:
                    questions.append({'name': parts[0].strip(), 'score': parts[1].strip()})
                elif len(parts) == 1:
                    questions.append({'name': f"题{len(questions) + 1}", 'score': parts[0].strip()})
        except:
            questions = [{'name': '一', 'score': '100'}]  # Fallback

        num_questions = len(questions)

        # 2. 获取学生成绩数据 (核心数据来源)
        # 根据 class_name 在数据库中查找 class_id，然后获取成绩
        students_data = []
        conn = db.get_connection()

        # 模糊匹配班级
        class_row = conn.execute("SELECT id FROM classes WHERE name LIKE ? LIMIT 1", (f"%{class_name}%",)).fetchone()
        if class_row:
            class_id = class_row['id']
            # 获取该班级所有学生及成绩
            db_students = db.get_students_with_grades(class_id)
            for stu in db_students:
                # 解析 score_details JSON 字符串
                # 假设 score_details 存的是 [{"name": "Task1", "score": 10}, ...] 或简单的 dict
                details = {}
                try:
                    raw_details = json.loads(stu['score_details']) if stu['score_details'] else []
                    # 如果是列表，转为字典方便按索引取
                    if isinstance(raw_details, list):
                        for idx, d in enumerate(raw_details):
                            # 尝试匹配题目顺序，这里简单按索引匹配
                            details[idx] = d.get('score', 0)
                    elif isinstance(raw_details, dict):
                        details = raw_details  # 暂不处理 dict 情况，按列表索引为准
                except:
                    pass

                students_data.append({
                    'student_id': stu['student_id'],
                    'name': stu['name'],
                    'total': stu['total_score'] if stu['total_score'] is not None else '',
                    'details': details
                })
        else:
            # 如果没找到班级，生成空表供老师手填，或者生成假数据
            # 这里生成 20 行空行
            for i in range(20):
                students_data.append({'student_id': '', 'name': '', 'total': '', 'details': {}})

        # 3. 页面布局设置
        # 题目要求的边距 (cm -> inch 转换由基类处理)
        self.setup_page_layout(top=1.4, bottom=1.9, left=1.3, right=1.3, header=0.8, footer=0.8)

        # 4. 计算列宽
        # 固定列：1(7), 2(15), 3(10), Last(11)
        # 剩余宽度：89 - 7 - 15 - 10 - 11 = 46
        # 中间列宽 = 46 / 题目数量
        fixed_width_sum = 7 + 15 + 10 + 11
        remain_width = 89 - fixed_width_sum
        mid_col_width = remain_width / num_questions if num_questions > 0 else 0

        self.set_col_width(1, 7)  # 序号
        self.set_col_width(2, 15)  # 学号
        self.set_col_width(3, 10)  # 姓名

        # 设置中间题目列宽
        start_q_col = 4
        for i in range(num_questions):
            self.set_col_width(start_q_col + i, mid_col_width)

        last_col_idx = start_q_col + num_questions
        self.set_col_width(last_col_idx, 11)  # 总分

        # 5. 绘制表头
        ws = self.ws

        # --- 第一行：大标题 ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_idx)
        cell_title = ws.cell(row=1, column=1, value="广西外国语学院机试（作品设计）考核登分表")
        self.set_row_height(1, 35)
        self.set_cell_style(cell_title, font_name=self.FONT_NAME_SONG, font_size=18, bold=True, border=False)

        # --- 第二行：课程信息 ---
        # 格式：课程：[Code]Name    专业年级班级：Class   (换行/空格)   授课老师：Name
        # 题目要求黑体12号居左，行高32
        info_text = f"课程：[{course_code}]{course_name}    专业年级班级：{class_name}\n授课老师：{teacher_name}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col_idx)
        cell_info = ws.cell(row=2, column=1, value=info_text)
        self.set_row_height(2, 32)
        # 注意：这里要求居左
        self.set_cell_style(cell_info, font_name=self.FONT_NAME_HEI, font_size=12, bold=False, align_h='left',
                            border=False)
        # 开启自动换行
        cell_info.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- 第三、四行：表头 (复杂合并) ---
        self.set_row_height(3, 15)
        self.set_row_height(4, 15)

        # 辅助函数：绘制合并表头并设置边框
        def draw_header_cell(r1, c1, r2, c2, text):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            cell = ws.cell(row=r1, column=c1, value=text)
            self.set_cell_style(cell, font_size=11, border=True)
            # 补全合并区域的边框 (openpyxl bug workaround)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    self.set_cell_style(ws.cell(row=r, column=c), border=True)
            return cell

        # 列1：序号 (合并 3-4 行)
        draw_header_cell(3, 1, 4, 1, "序号")

        # 列2：学号 (合并 3-4 行)
        draw_header_cell(3, 2, 4, 2, "学号")

        # 列3：姓名 (合并 3-4 行)
        draw_header_cell(3, 3, 4, 3, "姓名")

        # 中间列：题目 (Row 3 题号, Row 4 分数)
        for i, q in enumerate(questions):
            col = start_q_col + i
            # Row 3: 题号
            cell_q_name = ws.cell(row=3, column=col, value=q['name'])
            self.set_cell_style(cell_q_name, border=True)
            # Row 4: 分数
            cell_q_score = ws.cell(row=4, column=col, value=str(q['score']))
            self.set_cell_style(cell_q_score, border=True)

        # 最后一列：总分
        # Row 3: "总分"
        cell_total_label = ws.cell(row=3, column=last_col_idx, value="总分")
        self.set_cell_style(cell_total_label, border=True)
        # Row 4: 具体总分值 (由上层传入)
        cell_total_val = ws.cell(row=4, column=last_col_idx, value=str(total_score))
        self.set_cell_style(cell_total_val, border=True)

        # 6. 填充学生数据 (从第5行开始)
        current_row = 5
        for idx, student in enumerate(students_data):
            self.set_row_height(current_row, 18)

            # 序号
            c1 = ws.cell(row=current_row, column=1, value=idx + 1)
            self.set_cell_style(c1, border=True)

            # 学号
            c2 = ws.cell(row=current_row, column=2, value=student['student_id'])
            self.set_cell_style(c2, border=True)

            # 姓名
            c3 = ws.cell(row=current_row, column=3, value=student['name'])
            self.set_cell_style(c3, border=True)

            # 题目得分
            for i in range(num_questions):
                col = start_q_col + i
                # 从 student details 中获取分数，这里简单假设 details key 是 0,1,2...
                # 实际情况可能需要更复杂的映射
                score = student['details'].get(i, '')
                c_score = ws.cell(row=current_row, column=col, value=score)
                self.set_cell_style(c_score, border=True)

            # 总分
            c_total = ws.cell(row=current_row, column=last_col_idx, value=student['total'])
            self.set_cell_style(c_total, border=True)

            current_row += 1

        # 7. 保存文件
        self.wb.save(output_path)
        return output_path
