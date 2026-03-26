# -*- coding: utf-8 -*-
"""
Ergun River Narrative - SFL Spatial/Circumstance Pipeline (Customized Edition)
- Extract spatial circumstances (地点环境成分) based on customized SFL rules.
- Chunk by paragraphs, Call OpenAI-compatible API in 2 stages.
- Stage 1: Clause participant & basic places extraction.
- Stage 2: Deep spatial syntax & referent resolution.
- Aggregation: Computes 9 specific spatial metrics tables.
- Export to Excel.
"""

import os
import re
import json
import sys
import time
import hashlib
from dataclasses import dataclass, asdict
from typing import List, Dict, Any, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openai import OpenAI

# -----------------------------
# CONFIG
# -----------------------------
script_dir = os.path.dirname(os.path.abspath(__file__))
PDF_PATH = input("请输入PDF文件路径：").strip() or script_dir.strip('/') + "/doc/额尔古纳河右岸.pdf"
if not os.path.isfile(PDF_PATH):
    raise FileNotFoundError(f"PDF文件未找到：{PDF_PATH}")
OUT_DIR = os.path.dirname(PDF_PATH)

BASE_URL = "https://api.deepseek.com"
API_KEY = input("请输入OpenAI API Key：").strip()
MODEL = "deepseek-reasoner"

START_SENT = "我是雨和雪的老熟人了，我有九十岁了。"
END_SENT = "我落泪了，因为我已分不清天上人间了。"

CHUNK_MIN_CHARS = 500
CHUNK_TARGET_CHARS = 1000
CHUNK_MAX_CHARS = 1500
MAX_RETRIES = 4
RETRY_BACKOFF_SEC = 2.0
USE_CACHE = True
EXCEL_MAX_ROWS = 1_000_000

# -----------------------------
# CODING MANUAL / PROMPTS
# -----------------------------
PROMPT_STAGE_1 = r"""
你是一个“系统功能语言学（SFL）”及物性分析专家。
你需要对中文小说文本进行小句级的标注，核心任务是提取其中的【地点环境成分】（包括在名词短语中嵌入的小句也可以算作过程）。

【强制输出格式】
使用纯文本块，每个小句用 [RECORD_START] 和 [RECORD_END] 包裹。
切勿输出真正的 JSON！不要有引号、逗号或注释。严格遵守以下字段：

[RECORD_START]
clause_text: (原文句子)
human_participants: (句中所有人类参与者，如“人们,依芙琳”，用逗号分隔，无则填“无”)
nature_participants: (句中所有自然类参与者，如“黑熊,松树,驯鹿”，逗号分隔，无则填“无”)
basic_places: (提取句中的地点环境成分，如：在森林中, 神鼓上。用逗号分隔，无则填“无”)
[RECORD_END]
"""

PROMPT_STAGE_2 = r"""
你是一个空间语义与句法分析专家。
我将给你一组从小说中提取的句子和对应的【地点环境成分】。请你对该地点成分进行深度归类和句法分析。

【深度分类规则】
1. 代词消解：当地点成分是“那里/上面/里面”等代词时，必须根据上下文还原为它所指代的具体地点。
2. 主类型归类：自然类 / 人类身体部位类(如怀里) / 功能类(如商店中/神鼓上/房屋) / 地名类(如激流乡) / 方向类(如南边)。注意：如果方向词跟在其他地点后面（如“希愣柱的北侧”），请以中心词“希愣柱”为准，归类为“功能类”。
3. 自然子类型：如果主类型是自然类，细分为“自然无机物类”(如山、水) 或 “动物类”(如驯鹿的背上)。如果不是自然类，填“无”。
4. 中心词：提取该地点的中心词。
5. 树/林信息：若包含树或林，提取具体树种（如松树、杨树、桦树）和位置（如树梢上、树丛中）。
6. 方位词/介词：前,后,上,下,左,右,东,南,西,北,中,里,边,旁,外。
7. 宏观语法结构：如“介词+处所词+方位词”、“处所介词+处所词”、“趋向介词+处所词”、“趋向动词+处所词”、“处所词+方位词”等。
8. 微观处所词结构：如“名+名”、“形+名”、“名+的+名”、“形+形+的+名”、“数量词+形+的+名”、“动宾小句”、“动补小句”等。可根据实际语素关系准确修正。

【强制输出格式】
对于每个任务，严格使用以下纯文本格式：

[SPATIAL_START]
clause_text: (保持与原句一致)
circ_text: (保持与提取的地点一致)
resolved_target: (消解后的具体地点，如无代词则填原词的中心概念)
main_type: 自然类|人类身体部位类|功能类|地名类|方向类
nature_subtype: 自然无机物类|动物类|无
center_word: (中心词)
markers: (包含的方位词和介词，逗号分隔，无填无)
tree_species: (如桦树，无填无)
tree_position: (如树下，无填无)
macro_syntax: (如：介词+处所词+方位词)
micro_syntax: (如：名+的+名)
[SPATIAL_END]
"""


# -----------------------------
# Utilities
# -----------------------------
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def extract_pdf_text(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    pages = [doc.load_page(i).get_text("text") for i in range(doc.page_count)]
    return "\n".join(pages)


def slice_narrative_scope(full_text: str, start_sent: str, end_sent: str) -> str:
    compact = re.sub(r"[ \t\r\n\u3000]+", "", full_text)
    start_idx = compact.find(re.sub(r"[ \t\r\n\u3000]+", "", start_sent))
    if start_idx < 0: raise ValueError("未找到正文起点锚点。")
    end_compact = re.sub(r"[ \t\r\n\u3000]+", "", end_sent)
    end_idx = compact.find(end_compact)
    if end_idx < 0:
        m = re.search(r"我落泪了，?因为我已.*?分不清.*?天上人间了。", compact)
        if not m: raise ValueError("未找到正文终点锚点。")
        end_idx = m.end()
    else:
        end_idx += len(end_compact)
    return compact[start_idx:end_idx]


def sentence_split(text: str) -> List[str]:
    sents, buf = [], ""
    for ch in text:
        buf += ch
        if ch in "。！？；":
            t = buf.strip()
            if t: sents.append(t)
            buf = ""
    if buf.strip(): sents.append(buf.strip())
    return sents


def paragraph_split_from_sentences(sentences: List[str]) -> List[str]:
    paras, buf = [], []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            paras.append("".join(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf: paras.append("".join(buf))
    return [p for p in paras if p.strip()]


@dataclass
class Chunk:
    chunk_id: int
    start_sent_id: int
    end_sent_id: int
    char_count: int
    text_with_markers: str
    sha1: str


def build_chunks(sentences: List[str]) -> List[Chunk]:
    para_sent_counts, buf = [], []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            para_sent_counts.append(len(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf: para_sent_counts.append(len(buf))

    para_boundaries, end = [], 0
    for c in para_sent_counts:
        end += c
        para_boundaries.append(end)

    chunks, chunk_id, chunk_start_sid, cur_sents, cur_chars = [], 1, 1, [], 0

    def flush_chunk(end_sid: int):
        nonlocal chunk_id, chunk_start_sid, cur_sents, cur_chars
        if not cur_sents: return
        lines = [f"〔S{chunk_start_sid + i:06d}〕{s}" for i, s in enumerate(cur_sents)]
        text_with_markers = "\n".join(lines)
        chunks.append(
            Chunk(chunk_id, chunk_start_sid, end_sid, cur_chars, text_with_markers, sha1_text(text_with_markers)))
        chunk_id += 1
        chunk_start_sid = end_sid + 1
        cur_sents, cur_chars = [], 0

    for sid, s in enumerate(sentences, start=1):
        if cur_sents and (cur_chars + len(s) > CHUNK_MAX_CHARS): flush_chunk(sid - 1)
        cur_sents.append(s)
        cur_chars += len(s)
        if cur_chars >= CHUNK_TARGET_CHARS and (sid in para_boundaries):
            flush_chunk(sid)
        elif cur_chars >= CHUNK_MAX_CHARS:
            flush_chunk(sid)
    if cur_sents: flush_chunk(len(sentences))
    return chunks


def check_ai_connectivity_or_exit():
    try:
        requests.post(f"{BASE_URL.rstrip('/')}/chat/completions",
                      headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
                      json={"model": MODEL, "messages": [{"role": "user", "content": "say ok"}], "max_tokens": 5},
                      timeout=20).raise_for_status()
        print("  [Connect] API 连接成功")
    except SystemExit:
        raise
    except Exception as e:
        print(f"\n[ERROR] 接口访问失败: {e}")
        raise SystemExit(1)


def clean_ai_value(v: str) -> str:
    v = re.sub(r'//.*$', '', v)
    return re.sub(r'^["\']|["\']$', '', re.sub(r',$', '', v.strip())).strip()


def parse_stage1_text(raw_text: str, chunk_id: int) -> Dict[str, Any]:
    records, unparsed = [], []
    blocks = re.findall(r'\[RECORD_START\](.*?)\[RECORD_END\]', raw_text, re.DOTALL)
    if not blocks:
        blocks = [re.sub(r'\[RECORD_END\].*', '', b, flags=re.DOTALL).strip() for b in raw_text.split('[RECORD_START]')
                  if ":" in b or "：" in b]

    for block in blocks:
        record = {"chunk_id": chunk_id, "circumstances": []}
        places_raw = ""
        for line in block.strip().split('\n'):
            match = re.match(r'^"?([a-zA-Z_]+)"?\s*[:：]\s*(.*)$', line.strip())
            if match:
                k, v = match.groups()
                v = clean_ai_value(v)
                if v.lower() in ['null', 'none', '无', '-', '']: v = None
                if k == "basic_places":
                    places_raw = v
                else:
                    record[k] = v

        if record.get("clause_text"):
            if places_raw and places_raw not in ["无", "None"]:
                record["_pending_places"] = [p.strip() for p in places_raw.replace('，', ',').split(',') if p.strip()]
            records.append(record)
        else:
            unparsed.append({"raw_data": block[:50]})
    return {"chunk_id": chunk_id, "records": records, "unparsed": unparsed}


def parse_stage2_text(raw_text: str) -> Dict[str, Dict]:
    spatial_map = {}
    blocks = re.findall(r'\[SPATIAL_START\](.*?)\[SPATIAL_END\]', raw_text, re.DOTALL)
    for block in blocks:
        sd = {}
        for line in block.strip().split('\n'):
            match = re.match(r'^"?([a-zA-Z_]+)"?\s*[:：]\s*(.*)$', line.strip())
            if match:
                k, v = match.groups()
                v = clean_ai_value(v)
                if v.lower() in ['null', 'none', '无', '-', '']: v = None
                sd[k] = v
        if sd.get("clause_text") and sd.get("circ_text"):
            raw_markers = sd.get("markers", "")
            sd["markers"] = [m.strip() for m in
                             raw_markers.replace('，', ',').split(",")] if raw_markers and raw_markers != "无" else []
            spatial_map[f"{sd['clause_text']}_{sd['circ_text']}"] = sd
    return spatial_map


def call_ai_for_chunk(chunk: Chunk, cache_dir: str) -> Dict[str, Any]:
    ensure_dir(cache_dir)
    cache_path = os.path.join(cache_dir, f"chunk_{chunk.chunk_id:03d}_{chunk.sha1}.json")
    if USE_CACHE and os.path.exists(cache_path):
        print(f"  [chunk {chunk.chunk_id}] 读取缓存")
        with open(cache_path, "r", encoding="utf-8") as f: return json.load(f)

    def _call_api_stream(sys_prompt: str, usr_prompt: str, task_name: str) -> str:
        headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
        payload = {"model": MODEL,
                   "messages": [{"role": "system", "content": sys_prompt}, {"role": "user", "content": usr_prompt}],
                   "temperature": 0, "stream": True}
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                if attempt > 1: print(f"\n    {task_name} 重试 {attempt}/{MAX_RETRIES}... ", end="", flush=True)
                response = requests.post(f"{BASE_URL.rstrip('/')}/chat/completions", headers=headers, json=payload,
                                         timeout=(15, 600), stream=True)
                response.raise_for_status()
                full_content = []
                for line in response.iter_lines():
                    if not line: continue
                    line_text = line.decode("utf-8")
                    if line_text.startswith("data: ") and "[DONE]" not in line_text:
                        try:
                            data_json = json.loads(line_text[6:])
                            delta = data_json["choices"][0].get("delta", {})
                            piece = delta.get("content") or delta.get("reasoning_content")
                            if piece: full_content.append(piece)
                        except Exception:
                            pass
                print(" 完成")
                return re.sub(r"<think>.*?</think>", "", "".join(full_content).strip(), flags=re.DOTALL).strip()
            except Exception as e:
                print(f" [出错: {str(e)[:50]}]", end="", flush=True)
                time.sleep(RETRY_BACKOFF_SEC)
        raise RuntimeError(f"{task_name} 重试耗尽")

    print(f"    处理 chunk {chunk.chunk_id} ", end="", flush=True)
    s1_prompt = f"{PROMPT_STAGE_1}\n\n【分析文本】\n{chunk.text_with_markers}"
    parsed_stage1 = parse_stage1_text(_call_api_stream("严格遵循格式，禁止输出JSON符号。", s1_prompt, "Stage-1"),
                                      chunk.chunk_id)

    places_to_analyze = []
    for rec in parsed_stage1["records"]:
        if "_pending_places" in rec:
            for p in rec["_pending_places"]:
                places_to_analyze.append(f"- 原句：{rec.get('clause_text')}\n  地点：{p}")

    spatial_map = {}
    if places_to_analyze:
        print(f"      -> 发现 {len(places_to_analyze)} 个地点成分，启动深度解析 ", end="", flush=True)
        spatial_map = parse_stage2_text(_call_api_stream("专注空间句法分析。",
                                                         f"{PROMPT_STAGE_2}\n\n【需分析的句子与地点列表】\n" + "\n".join(
                                                             places_to_analyze), "Stage-2"))

    for rec in parsed_stage1["records"]:
        pending = rec.pop("_pending_places", [])
        for p in pending:
            sd = spatial_map.get(f"{rec.get('clause_text')}_{p}")
            rec["circumstances"].append({"type": "Place", "text": p, "spatial_details": sd})

    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(parsed_stage1, f, ensure_ascii=False, indent=2)
    return parsed_stage1


def flatten_records(all_payloads: List[Dict[str, Any]], chunk_map: Dict[int, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    recs, unparsed = [], []
    for payload in all_payloads:
        cid = payload.get("chunk_id", -1)
        if cid == -1 and payload.get("records"): cid = payload["records"][0].get("chunk_id", -1)
        ch = chunk_map.get(cid)
        for r in payload.get("records", []):
            if isinstance(r, dict):
                rec = dict(r)
                rec["chunk_id"] = cid
                rec["chunk_start_sent_id"] = ch.start_sent_id if ch else None
                rec["chunk_end_sent_id"] = ch.end_sent_id if ch else None
                recs.append(rec)
        for u in payload.get("unparsed", []):
            item = dict(u) if isinstance(u, dict) else {"raw_data": str(u), "error_reason": "Format Error"}
            item["chunk_id"] = cid
            unparsed.append(item)
    return pd.DataFrame(recs), pd.DataFrame(unparsed)


# -----------------------------
# Aggregation & Stats (The 9 Core Tables) - 学术严密修正终极版
# -----------------------------
def compute_spatial_tables(records_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    spatial_list = []
    nature_clause_counted = set()
    human_p_count = 0
    nature_p_count = 0

    # 辅助函数：安全读取值，消灭导致 Pandas 丢数据的 None/空值
    def safe_val(val, default="无"):
        if val is None: return default
        s = str(val).strip()
        if s.lower() in ["", "none", "null", "无", "-"]: return default
        return s

    # 1. 遍历并执行强力数据清洗与层级重构
    for idx, row in records_df.iterrows():
        circs = row.get("circumstances", [])
        clause_id = f"{row.get('chunk_id')}_{idx}"

        h_parts = row.get("human_participants")
        n_parts = row.get("nature_participants")

        has_nature_spatial = False

        if isinstance(circs, list) and len(circs) > 0:
            for c in circs:
                if isinstance(c, dict) and c.get("spatial_details"):
                    sd = c["spatial_details"]

                    # 提前安全提取所有需要查验的字段
                    raw_main = str(sd.get("main_type", "未分类"))
                    raw_sub = str(sd.get("nature_subtype", "无"))
                    text_val = str(c.get("text", ""))
                    target_val = safe_val(sd.get("resolved_target"), "未知")
                    macro_syn_val = safe_val(sd.get("macro_syntax"), "未识别")
                    micro_syn_val = safe_val(sd.get("micro_syntax"), "未识别")

                    # ==========================================
                    # 【调整1 - 终极版】彻底剔除时间环境成分
                    # 扫描覆盖：主类、子类、原词、消解词、宏观语法、微观语法
                    # ==========================================
                    if any("时间" in val for val in
                           [raw_main, raw_sub, text_val, target_val, macro_syn_val, micro_syn_val]):
                        continue  # 只要发现任何包含“时间”的痕迹，直接丢弃，不参与后续统计

                    # ==========================================
                    # 【调整3 & 4】主类型重构与地名类数据召回
                    # ==========================================
                    main_type = "未分类"
                    nature_subtype = "无"

                    # (a) 抽象类判定 (如: 世界, 梦里)
                    if "抽象" in raw_main or "世界" in target_val or "世界" in text_val:
                        main_type = "抽象类"
                    # (b) 地名类判定 (强制剥离括号后缀，找回丢失的数据)
                    elif "地名" in raw_main or "地名" in raw_sub:
                        main_type = "地名类"
                    # (c) 功能类判定
                    elif "功能" in raw_main:
                        main_type = "功能类"
                    # (d) 方向类保留
                    elif "方向" in raw_main or "方位" in raw_main:
                        main_type = "方向类"
                    # (e) 将“人类身体部位类”降级为“自然类”的子类
                    elif "身体" in raw_main or "部位" in raw_main:
                        main_type = "自然类"
                        nature_subtype = "人类身体部位类"
                    # (f) 剩余全部兜底为自然类
                    else:
                        main_type = "自然类"

                    # ==========================================
                    # 【调整2】自然类内部子类型精准重构 (植物/动物/无机物)
                    # ==========================================
                    if main_type == "自然类" and nature_subtype == "无":
                        # 强制植物类判定
                        if "植物" in raw_sub or any(kw in target_val or kw in text_val for kw in
                                                    ["树", "林", "草", "木", "花", "叶", "桦", "松", "柳", "苔藓",
                                                     "蘑菇", "森林"]):
                            nature_subtype = "植物类"
                        # 强制动物类判定
                        elif "动物" in raw_sub or any(kw in target_val or kw in text_val for kw in
                                                      ["驯鹿", "熊", "狗", "马", "狼", "鸟", "鱼", "鹰", "背上", "身上",
                                                       "角上"]):
                            nature_subtype = "动物类"
                        # 强制身体部位判定 (防漏)
                        elif "身体" in raw_sub or "部位" in raw_sub or any(kw in target_val or kw in text_val for kw in
                                                                           ["怀", "眼", "心", "手", "脚", "头", "脸",
                                                                            "肩"]):
                            nature_subtype = "人类身体部位类"
                        # 以上都不满足，才是自然无机物类
                        else:
                            nature_subtype = "自然无机物类"

                    # 规范化：非自然类必须没有子类型
                    if main_type != "自然类":
                        nature_subtype = "无"

                    # 标记小句是否含有自然类地点，供需求9统计参与者使用
                    if main_type == "自然类":
                        has_nature_spatial = True

                    # 最终安全入库
                    spatial_list.append({
                        "clause_text": str(row.get("clause_text", "")),
                        "text": text_val,
                        "resolved_target": target_val,
                        "main_type": main_type,
                        "nature_subtype": nature_subtype,
                        "center_word": safe_val(sd.get("center_word"), "未知"),
                        "markers": sd.get("markers", []),
                        "tree_species": safe_val(sd.get("tree_species"), "无"),
                        "tree_position": safe_val(sd.get("tree_position"), "无"),
                        "macro_syntax": macro_syn_val,  # 直接使用上方已安全提取的值
                        "micro_syntax": micro_syn_val,  # 直接使用上方已安全提取的值
                    })

        # 【需求9】统计参与者 (仅在含有自然类地点的小句中触发)
        if has_nature_spatial and clause_id not in nature_clause_counted:
            nature_clause_counted.add(clause_id)
            if h_parts and str(h_parts).strip() not in ["无", "None", ""]:
                human_p_count += len([p for p in str(h_parts).replace('，', ',').split(',') if p.strip()])
            if n_parts and str(n_parts).strip() not in ["无", "None", ""]:
                nature_p_count += len([p for p in str(n_parts).replace('，', ',').split(',') if p.strip()])

    # --- 聚合统计部分 ---
    spatial_df = pd.DataFrame(spatial_list)
    tables = {}
    if spatial_df.empty:
        return {"1_Total_Count": pd.DataFrame([{"Total_Spatial_Components": 0}])}

    # 现在的 total_spatial 是连根拔除所有含“时间”二字的数据后的最纯净基数
    total_spatial = len(spatial_df)

    # Req 1: Total Count
    tables["1_Total_Count"] = pd.DataFrame([{"Total_Spatial_Components": total_spatial}])

    # Req 2: Types, Subtypes, Ratios
    type_counts = spatial_df.groupby(["main_type", "nature_subtype"], dropna=False).size().reset_index(name="Count")
    type_counts["Ratio"] = type_counts["Count"] / total_spatial
    tables["2_Types_Subtypes"] = type_counts.sort_values(by=["Count"], ascending=False)

    nature_df = spatial_df[spatial_df["main_type"] == "自然类"].copy()
    total_nature = len(nature_df)

    # Req 3: Nature Keywords
    keywords = ["河", "山", "树", "地", "林", "水"]
    kw_stats = []
    if total_nature > 0:
        for kw in keywords:
            count = nature_df.apply(lambda x: kw in str(x["resolved_target"]) or kw in str(x["text"]), axis=1).sum()
            kw_stats.append({"Keyword": kw, "Count": count, "Ratio": count / total_nature})
    tables["3_Nature_Keywords"] = pd.DataFrame(kw_stats)

    # Req 4: Top 10 Nature locations
    if total_nature > 0:
        top10 = nature_df["resolved_target"].value_counts().head(10).reset_index()
        top10.columns = ["Location", "Count"]
        top10["Ratio"] = top10["Count"] / total_nature
        tables["4_Nature_Top10"] = top10

    # Req 5: Markers
    target_markers = ["前", "后", "上", "下", "左", "右", "东", "南", "西", "北", "中", "里", "边", "旁", "外"]
    marker_counts = {m: 0 for m in target_markers}
    for markers in spatial_df["markers"]:
        if isinstance(markers, list):
            for m in markers:
                for tm in target_markers:
                    if tm in str(m): marker_counts[tm] += 1
    m_df = pd.DataFrame(list(marker_counts.items()), columns=["Marker", "Count"])
    m_df["Ratio"] = m_df["Count"] / total_spatial
    tables["5_Markers_Dist"] = m_df.sort_values(by="Count", ascending=False)

    # Req 6: Trees
    is_tree_condition = nature_df["resolved_target"].astype(str).str.contains("树|林", regex=True) | nature_df[
        "text"].astype(str).str.contains("树|林", regex=True)
    tree_df = nature_df[is_tree_condition]
    total_trees = len(tree_df)
    if total_trees > 0:
        species_df = tree_df["tree_species"].replace("无", pd.NA).dropna().value_counts().reset_index()
        species_df.columns = ["Tree_Species", "Count"]
        species_df["Ratio"] = species_df["Count"] / total_trees
        tables["6_Tree_Species"] = species_df

        pos_df = tree_df["tree_position"].replace("无", pd.NA).dropna().value_counts().reset_index()
        pos_df.columns = ["Tree_Position", "Count"]
        pos_df["Ratio"] = pos_df["Count"] / total_trees
        tables["6_Tree_Positions"] = pos_df

    # Req 7: Macro Syntax (Nature only)
    if total_nature > 0:
        macro_syn = nature_df["macro_syntax"].value_counts().reset_index()
        macro_syn.columns = ["Macro_Syntax", "Count"]
        macro_syn["Ratio"] = macro_syn["Count"] / total_nature
        tables["7_Nature_MacroSyntax"] = macro_syn

    # Req 8: Micro Syntax (All)
    micro_syn = spatial_df["micro_syntax"].value_counts().reset_index()
    micro_syn.columns = ["Micro_Syntax", "Count"]
    micro_syn["Ratio"] = micro_syn["Count"] / total_spatial
    tables["8_All_MicroSyntax"] = micro_syn

    # Req 9: Participants
    total_p = human_p_count + nature_p_count
    p_data = [
        {"Participant_Type": "人类 (Human)", "Count": human_p_count,
         "Ratio": human_p_count / total_p if total_p else 0},
        {"Participant_Type": "自然类 (Nature)", "Count": nature_p_count,
         "Ratio": nature_p_count / total_p if total_p else 0}
    ]
    tables["9_Participants"] = pd.DataFrame(p_data)

    tables["Raw_Spatial_Data"] = spatial_df

    return tables


# -----------------------------
# Final AI Synthesis
# -----------------------------
def call_ai_for_final_synthesis(tables: Dict[str, pd.DataFrame], out_dir: str) -> Dict[str, Any]:
    def df_to_markdown(df: pd.DataFrame, max_rows=30) -> str:
        try:
            return df.head(max_rows).to_markdown(index=False)
        except ImportError:
            # 如果没有安装 tabulate，则降级使用纯文本字符串
            return df.head(max_rows).to_string(index=False)

    payload_md = "\n".join([f"## {name}\n{df_to_markdown(df)}\n" for name, df in tables.items() if not df.empty])
    system_msg = "You are a senior spatial narrative analyst. Output ONLY JSON."
    user_msg = (
        "You will receive aggregated statistics regarding the SPATIAL and ENVIRONMENTAL components of a Chinese eco-narrative.\n"
        "Tasks:\n"
        "1) Provide a concise narrative synthesis (Chinese) of how the text utilizes natural spaces, directional markers, and syntactic structures to convey spatial relationships.\n"
        "2) Summarize what the data says about human vs nature interaction in these spaces.\n"
        "Return STRICT JSON with keys:\n"
        "{\n"
        '  "spatial_narrative_summary_cn": str,\n'
        '  "human_nature_interaction_cn": str\n'
        "}\n\n"
        f"Aggregated tables:\n{payload_md}"
    )
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    payload = {"model": MODEL,
               "messages": [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
               "temperature": 0.2}
    resp = requests.post(f"{BASE_URL.rstrip('/')}/chat/completions", headers=headers, json=payload, timeout=60)
    data = json.loads(
        re.sub(r"^```json\s*", "", resp.json()["choices"][0]["message"]["content"].strip()).replace("```", ""))
    with open(os.path.join(out_dir, "final_ai_synthesis.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return data


# -----------------------------
# Excel Writer
# -----------------------------
def add_sheet_from_df(wb: Workbook, name: str, df: pd.DataFrame, percent_cols: Optional[List[str]] = None,
                      description: str = ""):
    ws = wb.create_sheet(title=name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) if not df.empty else 1)
    cell_desc = ws.cell(row=1, column=1, value=description)
    cell_desc.font, cell_desc.fill, cell_desc.alignment = Font(bold=True, italic=True, color="333333"), PatternFill(
        "solid", fgColor="E0E0E0"), Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    rows = list(dataframe_to_rows(df, index=False, header=True))
    if not rows: return
    for col_idx, value in enumerate(rows[0], start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font, cell.fill, cell.alignment = Font(bold=True), PatternFill("solid", fgColor="F2F2F2"), Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row_data in enumerate(rows[1:], start=3):
        for c_idx, value in enumerate(row_data, start=1): ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A3"
    for col_i, col_name in enumerate(df.columns, start=1):
        sample = df.iloc[:, col_i - 1].astype(str).head(50).tolist()
        ws.column_dimensions[get_column_letter(col_i)].width = min(
            max(10, (max([len(str(col_name))] + [len(x) for x in sample]) if sample else len(str(col_name))) * 1.2), 60)
        if percent_cols and col_name in percent_cols:
            for r in range(3, ws.max_row + 1): ws.cell(row=r, column=col_i).number_format = "0.00%"


def write_final_excel(out_path: str, tables: Dict[str, pd.DataFrame], final_ai: Optional[Dict[str, Any]] = None):
    print("正在写入 Excel ...")

    # 核心修复：序列化复杂列，防止含有 list 或 dict 导致 Excel 报错
    def serialize_complex_columns(df: pd.DataFrame):
        if df.empty: return df
        for col in df.columns:
            if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
                df[col] = df[col].apply(
                    lambda x: ",".join(str(i) for i in x) if isinstance(x, list) else (
                        json.dumps(x, ensure_ascii=False) if isinstance(x, dict) else x
                    )
                )
        return df

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    lines = [
        "统计分析报告：生态叙事空间与地点成分分析",
        "------------------------------------------------",
        "本报告专注于挖掘小说中所有的地点环境成分，及其背后的微观、宏观句法结构。",
        "所有数据根据指定要求的 9 项统计逻辑严格生成。"
    ]
    for i, line in enumerate(lines, start=1): ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    ws["A1"].font = Font(bold=True, size=14)

    table_descriptions = {
        "1_Total_Count": "【需求1】地点环境成分的总数量",
        "2_Types_Subtypes": "【需求2】地点环境成分的类型、子类型、数量、比例",
        "3_Nature_Keywords": "【需求3】自然类地点中 河、山、树、地、林、水的数量和比例",
        "4_Nature_Top10": "【需求4】自然类地点中排名前10的地点的数量和比例",
        "5_Markers_Dist": "【需求5】地点环境成分中，方位词和介词的数量和比例",
        "6_Tree_Species": "【需求6】自然地点环境成分中，树的种类的数量和比例",
        "6_Tree_Positions": "【需求6】自然地点环境成分中，树的不同位置的数量和比例",
        "7_Nature_MacroSyntax": "【需求7】自然类地点的语法结构的类型、数量和比例 (如 介词+处所词+方位词)",
        "8_All_MicroSyntax": "【需求8】处所词的语法结构类型以及数量、比例 (如 名+的+名)",
        "9_Participants": "【需求9】与自然类地点环境成分同时搭配出现的参与者中，人类与自然类参与者的数量和比例",
        "Raw_Spatial_Data": "【原始数据】所有提取并归类的地点成分明细"
    }

    percent_map = {k: ["Ratio"] for k in table_descriptions.keys()}

    for name, df in tables.items():
        if df.empty: continue
        # 在传递给 openpyxl 前，先清洗一遍数据
        safe_df = serialize_complex_columns(df.copy())
        add_sheet_from_df(wb, name, safe_df, percent_cols=percent_map.get(name, []),
                          description=table_descriptions.get(name, ""))

    if final_ai:
        ws2 = wb.create_sheet("AI_Synthesis_Report")
        ws2.column_dimensions["A"].width = 120
        ws2["A1"] = "空间叙事 AI 定性分析报告"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2["A3"], ws2["A4"] = "【空间运用与句法综述】", final_ai.get("spatial_narrative_summary_cn", "")
        ws2["A3"].font = Font(bold=True)
        ws2["A6"], ws2["A7"] = "【人与自然空间互动】", final_ai.get("human_nature_interaction_cn", "")
        ws2["A6"].font = Font(bold=True)

    wb.save(out_path)


# -----------------------------
# MAIN
# -----------------------------
def main():
    if not API_KEY: raise ValueError("API_KEY 为空。请在脚本顶部填入你的 key。")
    ensure_dir(OUT_DIR)
    cache_dir = os.path.join(OUT_DIR, "cache_spatial_chunks")

    print("[1] 提取PDF文本并切分正文...")
    sentences = sentence_split(slice_narrative_scope(extract_pdf_text(PDF_PATH), START_SENT, END_SENT))
    chunks = build_chunks(sentences)
    chunk_map = {c.chunk_id: c for c in chunks}

    check_ai_connectivity_or_exit()

    print("[2] 启动两阶段大模型解析...")
    payloads = []
    for ch in chunks:
        try:
            payloads.append(call_ai_for_chunk(ch, cache_dir))
        except Exception as e:
            print(f"  chunk {ch.chunk_id} 失败: {e}")

    print("[3] 展平记录并进行 9 项核心指标计算...")
    records_df, _ = flatten_records(payloads, chunk_map)
    tables = compute_spatial_tables(records_df)

    print("[4] AI 定性综合分析...")
    final_ai = None
    try:
        final_ai = call_ai_for_final_synthesis(tables, OUT_DIR)
    except Exception as e:
        print(f"WARN: final synthesis failed: {e}")

    print("[5] 写入 Excel...")
    out_xlsx = os.path.join(OUT_DIR, "Ergun_Spatial_Circumstances_Analysis-V4.xlsx")
    write_final_excel(out_xlsx, tables, final_ai=final_ai)
    print(f"DONE. 报告已保存至: {out_xlsx}")


if __name__ == "__main__":
    main()