# -*- coding: utf-8 -*-
"""
Ergun River Narrative - SFL Transitivity Pipeline
- Extract narrative text from PDF (skip author bio via anchors)
- Chunk by paragraphs (fallback to sentences), ~3000-5000 Chinese chars
- Call OpenAI-compatible Chat Completions API (multiple independent tasks) with a stable coding manual
- Validate/repair JSON, retry failures, resume from cache
- Aggregate records, compute required stats, ask AI for final consistency + narrative synthesis
- Export final Excel to local disk

Dependencies:
  pip install pymupdf pandas openpyxl
"""

import os
import re
import json
import sys
import time
import math
import hashlib
from dataclasses import dataclass, asdict
from typing import List, Dict, Any, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ====== OpenAI SDK (OpenAI-compatible base_url) ======
from openai import OpenAI

# -----------------------------
# CONFIG
# -----------------------------
PDF_PATH = input("请输入PDF文件路径：").strip()  # 输入PDF路径
if not os.path.isfile(PDF_PATH):
    raise FileNotFoundError(f"PDF文件未找到：{PDF_PATH}")
OUT_DIR = os.path.dirname(PDF_PATH)   # 输出目录

# BASE_URL = "https://xiaoai.plus/v1/"
# BASE_URL = "http://127.0.0.1:23333/v1/"
# BASE_URL = "https://api.mttieeo.com/v1/"
BASE_URL = "https://api.vectorengine.ai"
API_KEY = input("请输入OpenAI API Key：").strip()  # 输入API Key
# API_KEY = "cs-sk-f0fd0228-4540-470f-8237-789684ac5f7e"
# MODEL = input("输入模型ID（如 gpt-4-turbo-0613）：").strip() or "gpt-4o"
# MODEL = "[满血]gemini-3.0-pro-preview"
# MODEL = "gemini-2.5-pro-thinking"
MODEL = "gpt-5.1-thinking-all"

# Narrative anchors
START_SENT = "我是雨和雪的老熟人了，我有九十岁了。"
END_SENT = "我落泪了，因为我已分不清天上人间了。"

# Chunk sizing (chars, after whitespace normalization; Chinese chars ~ tokens close to 1:1)
CHUNK_MIN_CHARS = 500
CHUNK_TARGET_CHARS = 1000
CHUNK_MAX_CHARS = 1500

# Retry policy
MAX_RETRIES = 4
RETRY_BACKOFF_SEC = 2.0

# Cache policy (resume)
USE_CACHE = True

# Excel sheet limits
EXCEL_MAX_ROWS = 1_000_000  # xlsx limit ~1,048,576

# -----------------------------
# CODING MANUAL / PROMPT
# -----------------------------
CODING_MANUAL = r"""
你是一个“系统功能语言学（SFL）及物性系统”标注员，需要对中文小说叙事文本进行可复核的结构化标注。
你必须在多次独立任务中保持一致口径：所有规则以本手册为准。

【统计范围】
- 输入文本已来自小说正文（不含作者简介）。无需再处理作者简介。

【计数单位】
- 每一个“过程实例”（process instance）计1条记录：以“小句（clause）”为主，包含嵌入小句/嵌入短语中的过程（embedded process）。
- 小句切分参考：逗号/分号/冒号等；纯环境片段并入最近带过程核的小句。

【过程类型（固定6类）】
- Material（物质过程）：做/发生/对具体客体的处置
- Behavioural（行为过程）：身体化行为、感知生理表现
- Mental（心理过程）：感知/认知/情感/意愿
- Verbal（言语过程）：说/问/告诉/叫
- Relational（关系过程）：是/像/成为/属性描写/占有
- Existential（存在过程）：（地点）有/出现/存在X

【新增句法特征（用于深层语义分析）】
- is_ba_sentence (布尔值): 识别该小句是否使用“把”或“将”引出受事。
- is_bei_sentence (布尔值): 识别该小句是否使用“被”、“叫”、“让”、“给”等被动标记。
- copula_type (字符串): 仅限关系过程（Relational），根据文本提取系词。可选值为："是", "像", "似", "无" (不使用系词，如"山猫很狡猾")。非关系过程填 null。

【关键判别（已确认）】
- 被字句：若大自然/动物作主语且为受事（X被…），则其角色（Role）应标记为 "Goal"，同时 ActiveCounted=0。
- 存在过程：ActiveCounted=0

【主动参与者分类（ParticipantCategory）】
- Human：人类/人称代词/人物名
- HumanBodyPart：明确为“人类身体部位”
- Artifact：人造物
- Nature：除上之外一律归 Nature（包含动物部位及自然相关物）
  - NatureSubcategory：Animal / Plant / Environment / NatureOther

【参与者角色（Role）】
- Material -> Actor (动作者) 或 Goal (目标/受事，如被字句主语)
- Behavioural -> Behaver
- Mental -> Senser
- Verbal -> Sayer
- Relational -> Carrier
- Existential -> Existent

【环境成分 Circumstances 分类】
- Extent: duration/distance/frequency
- Location: time/place
- Manner: means/quality/comparison/degree
- Cause / Contingency / Accompaniment / Role / Matter / Angle

【JSON Schema】输出严格有效JSON，不含任何解释。
{
  "chunk_id": int,
  "records": [
    {
      "sent_id": int,
      "clause_id": str,
      "clause_text": str,
      "process_type": "Material|Behavioural|Mental|Verbal|Relational|Existential",
      "process_subtype": str|null,
      "process_lexeme": str,
      "active_participant": str,
      "participant_category": "Human|HumanBodyPart|Artifact|Nature",
      "nature_subcategory": "Animal|Plant|Environment|NatureOther"|null,
      "canonical_flora_fauna": str|null,
      "role": "Actor|Goal|Behaver|Senser|Sayer|Carrier|Existent",
      "is_ba_sentence": bool,
      "is_bei_sentence": bool,
      "copula_type": "是|像|似|无"|null,
      "circumstances": [
        {"type": str, "subtype": str|null, "text": str}
      ],
      "embedded_processes": [
        {"text": str, "process_type": str, "process_lexeme": str}
      ],
      "active_counted": 0|1,
      "notes": str|null
    }
  ],
  "unparsed": [ ... ]
}
"""


# -----------------------------
# Utilities
# -----------------------------
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def normalize_text_keep_para(text: str) -> str:
    """
    Keep paragraph boundaries, but normalize intra-line wraps and whitespace.
    Strategy:
      - Convert fullwidth spaces to normal removal
      - Keep blank lines as paragraph breaks
      - Remove spaces/tabs; merge wrapped lines within a paragraph
    """
    text = text.replace("\u3000", "")
    # unify newlines
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # remove trailing spaces per line
    lines = [re.sub(r"[ \t]+", "", ln) for ln in text.split("\n")]
    # rebuild paragraphs: blank line separates paragraphs
    paras = []
    buf = []
    for ln in lines:
        if ln == "":
            if buf:
                paras.append("".join(buf))
                buf = []
        else:
            buf.append(ln)
    if buf:
        paras.append("".join(buf))
    # remove empty paras
    paras = [p for p in paras if p.strip()]
    return "\n\n".join(paras)


def extract_pdf_text(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    pages = [doc.load_page(i).get_text("text") for i in range(doc.page_count)]
    return "\n".join(pages)


def slice_narrative_scope(full_text: str, start_sent: str, end_sent: str) -> str:
    """
    Use anchors to slice narrative正文，避开作者简介。
    End anchor may have whitespace breaks; we use regex with loose spaces removal.
    """
    # For robust matching, remove spaces/newlines but keep punctuation for anchor-finding
    compact = re.sub(r"[ \t\r\n\u3000]+", "", full_text)

    start_idx = compact.find(re.sub(r"[ \t\r\n\u3000]+", "", start_sent))
    if start_idx < 0:
        raise ValueError("未找到正文起点锚点（START_SENT）。请核对锚点句是否与PDF文本一致。")

    # For end, build tolerant regex on compact text
    end_compact = re.sub(r"[ \t\r\n\u3000]+", "", end_sent)
    end_idx = compact.find(end_compact)
    if end_idx < 0:
        # fallback: regex around key phrase “我落泪了...天上人间了。”
        m = re.search(r"我落泪了，?因为我已.*?分不清.*?天上人间了。", compact)
        if not m:
            raise ValueError("未找到正文终点锚点（END_SENT）。请核对终止句。")
        end_idx = m.end()
    else:
        end_idx += len(end_compact)

    scope_compact = compact[start_idx:end_idx]

    # Now we need a paragraph-preserving scope. Best effort:
    # We locate start/end in the original text by searching start/end in compact, then map back is hard.
    # Practical solution: extract text again paragraph-preserving, then locate using compact find within compact version and just use compact scope.
    # We will re-inject paragraph boundaries later by sentence markers, so compact scope is acceptable for analysis.
    return scope_compact


def sentence_split(text: str) -> List[str]:
    """
    Split by sentence-ending punctuation. Keep punctuation at end.
    """
    sents = []
    buf = ""
    for ch in text:
        buf += ch
        if ch in "。！？；":
            t = buf.strip()
            if t:
                sents.append(t)
            buf = ""
    if buf.strip():
        sents.append(buf.strip())
    return sents


def paragraph_split_from_sentences(sentences: List[str]) -> List[str]:
    """
    Recreate paragraphs roughly: since compact text loses original paras,
    we group sentences into paragraphs by heuristics:
      - If a sentence ends with '。' keep grouping; start new paragraph at dialogue markers or scene breaks.
    For chunking, paragraph boundaries are "soft"; we can just group every N sentences.
    """
    # Heuristic: new paragraph if sentence contains "——" or starts with "“" and previous ended with "。"
    paras = []
    buf = []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            paras.append("".join(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf:
        paras.append("".join(buf))
    return [p for p in paras if p.strip()]


@dataclass
class Chunk:
    chunk_id: int
    start_sent_id: int
    end_sent_id: int
    char_count: int
    text_with_markers: str
    sha1: str


def build_chunks(sentences: List[str]) -> List[Chunk]:
    """
    Chunk by paragraphs (heuristic), fallback to sentence boundaries.
    Add sentence markers to stabilize AI output: 〔S000123〕
    """
    paras = paragraph_split_from_sentences(sentences)

    # Build a mapping: paragraph -> sentence id range by reconstructing from concatenation
    # We'll just chunk by sentences directly but try to cut at para boundaries by using paras list
    # and tracking sentence pointers.
    chunks: List[Chunk] = []
    sent_idx = 1
    chunk_id = 1
    # Build para sentence lists: distribute sentences into paras in same order
    # Since paras were built from sentences sequentially, we can reconstruct by re-splitting paras (not needed).
    # We'll instead chunk via sentences with target sizes and allow "soft" paragraph breaks at para ends.
    # Precompute para boundaries in sentence indices:
    para_boundaries = []
    acc = 0
    # We reconstruct by iterating paras and counting how many sentences were in each by greedy matching
    # because paras were built from sentences with join, we can store counts during paragraph_split, but we didn't.
    # We'll re-do paragraph_split with counts:
    para_sent_counts = []
    buf = []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            para_sent_counts.append(len(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf:
        para_sent_counts.append(len(buf))

    # Build para end indices
    end = 0
    for c in para_sent_counts:
        end += c
        para_boundaries.append(end)  # sentence index (1-based) where a paragraph ends

    def is_para_end(sid: int) -> bool:
        return sid in para_boundaries

    cur_sents = []
    cur_chars = 0
    chunk_start_sid = 1

    def flush_chunk(end_sid: int):
        nonlocal chunk_id, chunk_start_sid, cur_sents, cur_chars
        if not cur_sents:
            return
        # build text with sentence markers
        lines = []
        sid = chunk_start_sid
        for s in cur_sents:
            lines.append(f"〔S{sid:06d}〕{s}")
            sid += 1
        text_with_markers = "\n".join(lines)
        ch = Chunk(
            chunk_id=chunk_id,
            start_sent_id=chunk_start_sid,
            end_sent_id=end_sid,
            char_count=cur_chars,
            text_with_markers=text_with_markers,
            sha1=sha1_text(text_with_markers),
        )
        chunks.append(ch)
        chunk_id += 1
        chunk_start_sid = end_sid + 1
        cur_sents = []
        cur_chars = 0

    for sid, s in enumerate(sentences, start=1):
        s_len = len(s)
        # if adding would exceed max, flush before adding (ensure non-empty)
        if cur_sents and (cur_chars + s_len > CHUNK_MAX_CHARS):
            flush_chunk(sid - 1)

        cur_sents.append(s)
        cur_chars += s_len

        # flush when reaching target AND at paragraph end (preferred)
        if cur_chars >= CHUNK_TARGET_CHARS and is_para_end(sid):
            flush_chunk(sid)
        # or if very large, flush anyway at sentence end
        elif cur_chars >= CHUNK_MAX_CHARS:
            flush_chunk(sid)

    # remainder
    if cur_sents:
        flush_chunk(len(sentences))

    return chunks


# -----------------------------
# AI Call + JSON Validation
# -----------------------------
def make_client() -> OpenAI:
    # 显著增加 timeout，标注任务建议至少 120s，如果是慢速模型建议 300s
    # 同时增加 httpx 限制，防止连接过快断开
    import httpx
    http_client = httpx.Client(
        limits=httpx.Limits(max_connections=5, max_keepalive_connections=2),
        timeout=httpx.Timeout(300.0, read=240.0, connect=20.0)
    )
    client = OpenAI(
        base_url=BASE_URL,
        api_key=API_KEY,
        http_client=http_client
    )
    return client


def extract_visible_text_from_choice_message(msg) -> str:
    """
    Compatible with normal models (message.content) and reasoning models
    (message.reasoning_content). Return the best-effort visible text.
    """
    content = getattr(msg, "content", None)
    if isinstance(content, str) and content.strip():
        return content.strip()

    # some reasoning models put output here
    rc = getattr(msg, "reasoning_content", None)
    if isinstance(rc, str) and rc.strip():
        return rc.strip()

    # fallback: stringify
    try:
        s = str(msg)
        return s.strip()
    except Exception:
        return ""


def check_ai_connectivity_or_exit():
    """
    使用原生 requests 检查连接
    """
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": "say ok"}],
        "max_tokens": 5
    }
    try:
        # 使用较短的 timeout 进行测试
        response = requests.post(f"{BASE_URL.rstrip('/')}/chat/completions",
                                 headers=headers, json=payload, timeout=20)
        response.raise_for_status()
        print("  [Connect] API 连接成功")
    except Exception as e:
        print(f"\n[ERROR] 接口访问失败: {e}")
        if response := getattr(e, 'response', None):
            print(f"状态码: {response.status_code}, 返回内容: {response.text}")
        raise SystemExit(1)


ALLOWED_PROCESS_TYPES = {"Material", "Behavioural", "Mental", "Verbal", "Relational", "Existential"}
ALLOWED_PARTICIPANT_CATS = {"Human", "HumanBodyPart", "Artifact", "Nature"}
ALLOWED_ROLES = {"Actor", "Goal", "Behaver", "Senser", "Sayer", "Carrier", "Existent"}


def basic_validate_payload(payload: Dict[str, Any]) -> Tuple[bool, List[str]]:
    errs = []
    if not isinstance(payload, dict):
        return False, ["payload_not_dict"]
    if "chunk_id" not in payload or not isinstance(payload["chunk_id"], int):
        errs.append("missing_or_bad_chunk_id")
    if "records" not in payload or not isinstance(payload["records"], list):
        errs.append("missing_or_bad_records")

    for i, r in enumerate(payload.get("records", [])):
        if not isinstance(r, dict):
            errs.append(f"record_{i}_not_dict")
            continue
        # 移除了过于严苛的必填字段检查，因为有些模型偶尔会漏掉新增的可选布尔字段，防止重试风暴
        pt = r.get("process_type")
        if pt and pt not in ALLOWED_PROCESS_TYPES:
            errs.append(f"record_{i}_bad_process_type:{pt}")
        pc = r.get("participant_category")
        if pc and pc not in ALLOWED_PARTICIPANT_CATS:
            errs.append(f"record_{i}_bad_participant_category:{pc}")
        role = r.get("role")
        if role and role not in ALLOWED_ROLES:
            errs.append(f"record_{i}_bad_role:{role}")

    return (len(errs) == 0), errs


def is_cloudflare_524_or_timeout(err: Exception) -> bool:
    """
    Best-effort detection for Cloudflare 524 / timeouts from OpenAI-compatible SDKs.
    """
    s = str(err).lower()
    status = getattr(err, "status_code", None)
    if status == 524:
        return True
    if "error code 524" in s or "cloudflare" in s and "524" in s:
        return True
    if "timeout" in s or "timed out" in s or "readtimeout" in s:
        return True
    return False


def split_marked_text_by_lines(marked_text: str, sub_max_chars: int = 1800) -> List[str]:
    """
    marked_text is multiple lines like: 〔S000123〕....
    Split into smaller parts without breaking a line.
    """
    lines = [ln for ln in marked_text.split("\n") if ln.strip()]
    parts = []
    buf = []
    cur = 0
    for ln in lines:
        ln_len = len(ln)
        if buf and (cur + ln_len > sub_max_chars):
            parts.append("\n".join(buf))
            buf = [ln]
            cur = ln_len
        else:
            buf.append(ln)
            cur += ln_len
    if buf:
        parts.append("\n".join(buf))
    return parts


def call_ai_for_chunk(chunk: Chunk, cache_dir: str) -> Dict[str, Any]:
    """
    修复版：
    1. 移除 response_format 强制限制，兼容所有思考型(Thinking)模型。
    2. 增加对 <think> 标签的正则剥离。
    3. 捕获并暴露 API 在数据流中的隐藏报错。
    4. 用不同的符号区分“思考中 (*)”和“输出中 (.)”。
    """
    ensure_dir(cache_dir)
    cache_path = os.path.join(cache_dir, f"chunk_{chunk.chunk_id:03d}_{chunk.sha1}.json")

    # 缓存命中检查
    if USE_CACHE and os.path.exists(cache_path):
        with open(cache_path, "r", encoding="utf-8") as f:
            print(f"  [chunk {chunk.chunk_id}] 读取缓存: {os.path.basename(cache_path)}")
            return json.load(f)

    def _one_call(marked_text: str, current_chunk_id: int, sub_idx: int = 0) -> Dict[str, Any]:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "application/json, text/event-stream",
        }

        system_prompt = "You are a rigorous SFL transitivity annotator. Follow the coding manual exactly. Output ONLY valid JSON."
        user_content = (
            f"{CODING_MANUAL}\n\n"
            f"【Task Info: chunk_id={current_chunk_id}】\n"
            f"Please annotate the following text:\n{marked_text}\n\n"
            f"Output ONLY JSON. Do not include any text outside the JSON block."
        )

        payload = {
            "model": MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            "temperature": 0,
            "stream": True
            # 注意：这里去掉了 response_format，防止思考模型崩溃
        }

        last_err = None
        sub_mark = f" (分片-{sub_idx})" if sub_idx > 0 else ""
        print(f"    处理 chunk {current_chunk_id}{sub_mark} ", end="", flush=True)

        for attempt in range(1, MAX_RETRIES + 1):
            full_content = []
            try:
                if attempt > 1:
                    print(f"\n    重试 {attempt}/{MAX_RETRIES}... ", end="", flush=True)

                # 兼容部分 API 需要 /v1 前缀
                api_url = f"{BASE_URL.rstrip('/')}/chat/completions"
                if "/v1" not in api_url and "api.vectorengine.ai" in api_url:
                    api_url = f"{BASE_URL.rstrip('/')}/v1/chat/completions"

                response = requests.post(
                    api_url,
                    headers=headers,
                    json=payload,
                    timeout=(15, 600),
                    stream=True
                )

                if response.status_code != 200:
                    try:
                        err_msg = response.text[:200]
                    except:
                        err_msg = "Unknown"
                    raise RuntimeError(f"HTTP {response.status_code}: {err_msg}")

                char_counter = 0
                for line in response.iter_lines():
                    if not line:
                        continue

                    line_text = line.decode("utf-8")
                    if line_text.startswith("data: "):
                        data_str = line_text[6:]
                        if data_str.strip() == "[DONE]":
                            break

                        try:
                            data_json = json.loads(data_str)

                            # --- 关键修复 1：拦截流式隐藏报错 ---
                            if "error" in data_json:
                                raise RuntimeError(f"API流内报错: {data_json.get('error')}")

                            if not data_json.get("choices"):
                                continue

                            choice = data_json["choices"][0]
                            delta = choice.get("delta", {})

                            # --- 关键修复 2：兼容思考模型的字段 ---
                            content_piece = delta.get("content")
                            reasoning_piece = delta.get("reasoning_content")

                            if content_piece:  # 真正的 JSON 输出
                                full_content.append(content_piece)
                                char_counter += len(content_piece)
                                if char_counter > 20:
                                    sys.stdout.write(".")
                                    sys.stdout.flush()
                                    char_counter = 0
                            elif reasoning_piece:  # 模型的深度思考过程
                                full_content.append(reasoning_piece)  # 有些模型把思考也混在content里，统一收集后续正则清洗
                                char_counter += len(reasoning_piece)
                                if char_counter > 40:
                                    sys.stdout.write("*")  # 打印星号表示正在思考
                                    sys.stdout.flush()
                                    char_counter = 0

                        except (json.JSONDecodeError, IndexError, AttributeError):
                            continue

                print(" 完成")

                raw_content = "".join(full_content).strip()
                if not raw_content:
                    raise RuntimeError("流式响应内容为空，API可能拦截了请求。")

                # --- 关键修复 3：强力清洗数据，剔除 <think> 标签 ---
                # 剥离模型思考过程
                clean_content = re.sub(r"<think>.*?</think>", "", raw_content, flags=re.DOTALL).strip()

                # 提取 JSON 块
                clean_content = re.sub(r"^```json\s*", "", clean_content)
                clean_content = re.sub(r"\s*```$", "", clean_content)
                s = clean_content.find('{')
                e = clean_content.rfind('}')
                if s != -1 and e != -1:
                    clean_content = clean_content[s: e + 1]

                try:
                    parsed = json.loads(clean_content)
                except json.JSONDecodeError as je:
                    raise ValueError(f"JSON解析失败: {str(je)[:50]}")

                ok, errs = basic_validate_payload(parsed)
                if not ok:
                    raise ValueError(f"Schema校验失败: {errs[:3]}")

                return parsed

            except Exception as e:
                last_err = e
                print(f" [错误: {str(e)[:80]}]", end="", flush=True)
                time.sleep(RETRY_BACKOFF_SEC * attempt)

        print("")
        raise RuntimeError(f"重试耗尽，最后错误: {last_err}")

    # 主逻辑：先尝试整体，失败则拆分
    try:
        result = _one_call(chunk.text_with_markers, chunk.chunk_id)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        return result

    except Exception as e:
        print(f"\n    [chunk {chunk.chunk_id}] 整体失败，切换为子分片模式...")
        sub_texts = split_marked_text_by_lines(chunk.text_with_markers, sub_max_chars=600)
        all_records = []
        all_unparsed = []

        for i, sub_text in enumerate(sub_texts, start=1):
            try:
                sub_res = _one_call(sub_text, chunk.chunk_id, sub_idx=i)
                all_records.extend(sub_res.get("records", []))
                all_unparsed.extend(sub_res.get("unparsed", []))
            except Exception as sub_e:
                print(f"\n    [分片 {i} 失败] {sub_e}")
                all_unparsed.append({"sent_id": -1, "text": "SUB_CHUNK_FAILED", "reason": str(sub_e)})

        merged = {
            "chunk_id": chunk.chunk_id,
            "records": all_records,
            "unparsed": all_unparsed
        }

        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)

        return merged



def repair_json_via_ai(client: OpenAI, bad_text: str, errs: List[str]) -> Dict[str, Any]:
    """
    Ask AI to output valid JSON only, repairing prior output.
    """
    system_msg = "You are a strict JSON repair tool. Output ONLY valid JSON."
    user_msg = (
        "The previous output is not valid per schema. "
        f"Errors: {errs[:20]}\n"
        "Rewrite it into a STRICTLY valid JSON matching the schema. "
        "Do not add extra commentary.\n\n"
        f"Previous output:\n{bad_text}"
    )
    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.0,
    )
    content = resp.choices[0].message.content.strip()
    content = re.sub(r"^```json\s*", "", content)
    content = re.sub(r"\s*```$", "", content)
    return json.loads(content)


# -----------------------------
# Aggregation & Stats
# -----------------------------
def flatten_records(all_payloads: List[Dict[str, Any]], chunk_map: Dict[int, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    将嵌套的 JSON 展平为 DataFrame。
    [健壮性增强版]：防止大模型返回非字典格式导致程序崩溃 (dict() 转换错误)。
    """
    recs = []
    unparsed = []

    for payload in all_payloads:
        # 获取 chunk_id 兜底，防止 payload 格式错误
        cid = payload.get("chunk_id", -1)
        ch = chunk_map.get(cid)

        # 1. 处理主句记录 (Records)
        for r in payload.get("records", []):
            # [防御性检查 1] 防止模型把 records 写成了字符串列表
            if not isinstance(r, dict):
                unparsed.append({
                    "chunk_id": cid,
                    "raw_data": str(r),
                    "error_reason": "Record Format Error: Not a dictionary"
                })
                continue

            # 基础信息
            rec = dict(r)
            rec["chunk_id"] = cid
            rec["chunk_start_sent_id"] = ch.start_sent_id if ch else None
            rec["chunk_end_sent_id"] = ch.end_sent_id if ch else None
            rec["is_embedded"] = False  # 标记为主句

            # 确保关键字段存在
            if "active_counted" not in rec:
                rec["active_counted"] = 0

            recs.append(rec)

            # 2. 处理嵌入小句 (Embedded Processes)
            if "embedded_processes" in r and isinstance(r["embedded_processes"], list):
                for emb in r["embedded_processes"]:
                    # [防御性检查 2] 防止嵌入小句不是字典
                    if not isinstance(emb, dict):
                        continue

                    emb_rec = {
                        "chunk_id": cid,
                        "sent_id": rec.get("sent_id"),
                        "clause_id": f"{rec.get('clause_id', '')}_emb",
                        "clause_text": emb.get("text", "") or emb.get("clause_text", ""),
                        "process_type": emb.get("process_type"),
                        "process_lexeme": emb.get("process_lexeme"),
                        "active_participant": None,
                        "participant_category": None,
                        "role": None,
                        "active_counted": 0,
                        "is_embedded": True,
                        "notes": "Derived from embedded_processes"
                    }
                    recs.append(emb_rec)

        # 3. 处理解析失败的片段 (Unparsed)
        for u in payload.get("unparsed", []):
            # [防御性检查 3] 核心修复点：防止模型返回非字典格式（如字符串）
            if isinstance(u, dict):
                item = dict(u)
            else:
                # 如果模型返回了纯字符串或列表，把它包进一个安全的字典里
                item = {
                    "raw_data": str(u),
                    "error_reason": "Unparsed Format Error: Returned as string instead of dict"
                }

            item["chunk_id"] = cid
            unparsed.append(item)

    records_df = pd.DataFrame(recs)
    unparsed_df = pd.DataFrame(unparsed)
    return records_df, unparsed_df


def safe_pct(n: int, d: int) -> float:
    return (n / d) if d else 0.0


def compute_required_tables(records_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    if records_df.empty:
        raise ValueError("records_df is empty; no parsed records to summarize.")

    # 补全列名防止报错（包括新增的句法特征列）
    expected_cols = [
        "process_type", "participant_category", "nature_subcategory",
        "canonical_flora_fauna", "role", "active_counted", "is_embedded",
        "is_ba_sentence", "is_bei_sentence", "copula_type", "circumstances"
    ]
    for col in expected_cols:
        if col not in records_df.columns:
            # 布尔列默认 False，其他默认 None
            if col in ["is_ba_sentence", "is_bei_sentence"]:
                records_df[col] = False
            else:
                records_df[col] = None

    # ---- 基础表：前置原有逻辑 (已精简提取) ----
    proc = records_df[records_df["process_type"].notna()].groupby("process_type").size().reset_index(
        name="n").sort_values("n", ascending=False)
    proc["pct"] = proc["n"] / proc["n"].sum()

    active_df = records_df[records_df["active_counted"] == 1].copy()
    n_active = len(active_df)
    nature_all = active_df[active_df["participant_category"] == "Nature"].copy()
    n_nature = len(nature_all)

    nature_sub_sum = nature_all.groupby("nature_subcategory").size().reset_index(name="n").sort_values("n",
                                                                                                       ascending=False) if n_nature > 0 else pd.DataFrame(
        columns=["nature_subcategory", "n", "pct_of_nature"])
    if not nature_sub_sum.empty: nature_sub_sum["pct_of_nature"] = nature_sub_sum["n"] / n_nature

    cat_summary = pd.DataFrame([
        {"Category": "Nature", "n": len(active_df[active_df["participant_category"] == "Nature"])},
        {"Category": "Human (incl. BodyPart)",
         "n": len(active_df[active_df["participant_category"].isin(["Human", "HumanBodyPart"])])},
        {"Category": "Artifact", "n": len(active_df[active_df["participant_category"] == "Artifact"])}
    ])
    cat_summary["pct_of_active"] = cat_summary["n"] / n_active if n_active else 0.0

    role_sum = nature_all.groupby("role").size().reset_index(name="n").sort_values("n", ascending=False)
    role_sum["pct_of_nature_active"] = role_sum["n"] / n_nature if n_nature else 0.0

    nature_proc = nature_all.groupby("process_type").size().reset_index(name="n").sort_values("n", ascending=False)
    nature_proc["pct_of_nature_active"] = nature_proc["n"] / n_nature if n_nature else 0.0

    ff = nature_all[nature_all["canonical_flora_fauna"].notna()].copy()
    top10 = ff.groupby("canonical_flora_fauna").size().reset_index(name="n").sort_values("n", ascending=False).head(10)
    top10["pct_of_flora_fauna_active"] = top10["n"] / len(ff) if len(ff) else 0.0

    # (省略TopN详细代码以节省空间，直接使用原逻辑)
    TOPN_ANIMALS = 15
    animal_ff = ff[ff["nature_subcategory"] == "Animal"].copy()
    if not animal_ff.empty:
        animal_totals = animal_ff.groupby("canonical_flora_fauna").size().reset_index(name="Total").sort_values("Total",
                                                                                                                ascending=False)
        animal_totals["pct_of_all_animals"] = animal_totals["Total"] / animal_totals["Total"].sum()
        top_animals = animal_totals.head(TOPN_ANIMALS)["canonical_flora_fauna"].tolist()
        animal_top_df = animal_ff[animal_ff["canonical_flora_fauna"].isin(top_animals)]
        animal_role_counts = animal_top_df.pivot_table(index="canonical_flora_fauna", columns="role", values="sent_id",
                                                       aggfunc="count", fill_value=0).reset_index()
        for rc in ["Actor", "Behaver", "Senser", "Sayer", "Carrier"]:
            if rc not in animal_role_counts.columns: animal_role_counts[rc] = 0
        TopN_Animals_ByRole = animal_totals[animal_totals["canonical_flora_fauna"].isin(top_animals)].merge(
            animal_role_counts, on="canonical_flora_fauna", how="left").sort_values("Total", ascending=False)
    else:
        TopN_Animals_ByRole = pd.DataFrame(
            columns=["canonical_flora_fauna", "Total", "pct_of_all_animals", "Actor", "Behaver", "Senser", "Sayer",
                     "Carrier"])

    # ================= 新增分析模块 =================

    # ---- 新增 1 & 2: 物质过程中的把字句 (Actor) 和被字句 (Goal) ----
    nature_mat_actor = records_df[
        (records_df["process_type"] == "Material") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Actor")]
    ba_total = len(nature_mat_actor)
    ba_count = nature_mat_actor["is_ba_sentence"].sum() if ba_total > 0 else 0
    ba_stats = pd.DataFrame([{
        "Analysis_Type": "大自然作为物质过程动作者(Actor)",
        "Total_Instances": ba_total,
        "Target_Sentence_Type": "把字句/将字句",
        "Count": ba_count,
        "Ratio": ba_count / ba_total if ba_total > 0 else 0.0
    }])

    nature_mat_goal = records_df[
        (records_df["process_type"] == "Material") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Goal")]
    bei_total = len(nature_mat_goal)
    bei_count = nature_mat_goal["is_bei_sentence"].sum() if bei_total > 0 else 0
    bei_stats = pd.DataFrame([{
        "Analysis_Type": "大自然作为物质过程目标(Goal)",
        "Total_Instances": bei_total,
        "Target_Sentence_Type": "被动句式",
        "Count": bei_count,
        "Ratio": bei_count / bei_total if bei_total > 0 else 0.0
    }])
    syntax_ba_bei = pd.concat([ba_stats, bei_stats], ignore_index=True)

    # ---- 新增 3: 关系过程中的系词分布 ----
    nature_carrier = records_df[
        (records_df["process_type"] == "Relational") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Carrier")]
    carrier_total = len(nature_carrier)
    if carrier_total > 0:
        # 将空值填充为"未识别"
        nature_carrier.loc[:, "copula_type"] = nature_carrier["copula_type"].fillna("未识别")
        copula_stats = nature_carrier["copula_type"].value_counts().reset_index()
        copula_stats.columns = ["Copula_Type (系词类型)", "Count"]
        copula_stats["Ratio"] = copula_stats["Count"] / carrier_total
    else:
        copula_stats = pd.DataFrame(columns=["Copula_Type (系词类型)", "Count", "Ratio"])

    # ---- 新增 4: 环境成分 (Circumstances) 分布 ----
    circ_list = []
    for idx, row in records_df.iterrows():
        circs = row.get("circumstances", [])
        is_nature = (row.get("participant_category") == "Nature")
        # 宽泛匹配，哪怕 Nature 不是 Active，只要它出现在句子中我们就打标
        is_nature_active = is_nature and row.get("active_counted") == 1

        if isinstance(circs, list):
            for c in circs:
                if isinstance(c, dict):
                    ctype = str(c.get("type", "")).lower()
                    csubtype = str(c.get("subtype", "")).lower()

                    # 归一化为三大类
                    norm_type = "Other"
                    if "time" in ctype or "time" in csubtype or "extent" in ctype:
                        norm_type = "Time (时间/时长)"
                    elif "place" in ctype or "place" in csubtype or "location" in ctype:
                        norm_type = "Place (地点/空间)"
                    elif "manner" in ctype:
                        norm_type = "Manner (方式)"

                    circ_list.append({
                        "Circumstance_Type": norm_type,
                        "Is_Nature_Active": is_nature_active
                    })

    circ_df = pd.DataFrame(circ_list)
    if not circ_df.empty:
        circ_all = circ_df["Circumstance_Type"].value_counts().reset_index()
        circ_all.columns = ["Circumstance_Type", "Total_Count"]
        circ_all["Total_Ratio"] = circ_all["Total_Count"] / circ_all["Total_Count"].sum()

        circ_nature = circ_df[circ_df["Is_Nature_Active"]]["Circumstance_Type"].value_counts().reset_index()
        circ_nature.columns = ["Circumstance_Type", "Nature_Active_Count"]

        circ_stats = pd.merge(circ_all, circ_nature, on="Circumstance_Type", how="left").fillna(0)
        total_nature_circs = circ_stats["Nature_Active_Count"].sum()
        circ_stats["Nature_Active_Ratio"] = circ_stats[
                                                "Nature_Active_Count"] / total_nature_circs if total_nature_circs > 0 else 0.0
    else:
        circ_stats = pd.DataFrame(
            columns=["Circumstance_Type", "Total_Count", "Total_Ratio", "Nature_Active_Count", "Nature_Active_Ratio"])

    return {
        "Summary_ProcessTypes": proc,
        "Summary_NatureSubcategory": nature_sub_sum,
        "Summary_NatureVsHuman": cat_summary,
        "NatureActive_RoleDist": role_sum,
        "NatureActive_ProcessDist": nature_proc,
        "Top10_SpeciesActive": top10,
        "TopN_Animals_Detailed": TopN_Animals_ByRole,
        # 新增的论文报表
        "Syntax_Ba_Bei": syntax_ba_bei,
        "Syntax_Copula_Relational": copula_stats,
        "Circumstances_Dist": circ_stats
    }



# -----------------------------
# Final AI synthesis (optional but requested)
# -----------------------------
def call_ai_for_final_synthesis(client: OpenAI, tables: Dict[str, pd.DataFrame], out_dir: str) -> Dict[str, Any]:
    """
    Send summarized tables (not full records) back to AI for:
      - consistency check suggestions
      - narrative ecological linguistics insight
    Output JSON and save.
    """

    def df_to_markdown(df: pd.DataFrame, max_rows=50) -> str:
        return df.head(max_rows).to_markdown(index=False)

    payload_md = []
    for name, df in tables.items():
        payload_md.append(f"## {name}\n{df_to_markdown(df)}\n")
    summary_blob = "\n".join(payload_md)

    system_msg = "You are a senior SFL + ecolinguistics reviewer. Output ONLY JSON."
    user_msg = (
        "You will receive aggregated statistics from a clause-level SFL transitivity annotation of a Chinese eco-narrative.\n"
        "Tasks:\n"
        "1) Check internal consistency and flag likely coding inconsistencies to audit (e.g., 有 existential vs possessive; 看/看见; 被字句 active_counted).\n"
        "2) Provide a concise narrative synthesis (Chinese) of what the distributions imply for 'nature agency' in the text.\n"
        "3) Provide concrete next-step QA suggestions (what to spot-check).\n"
        "Return STRICT JSON with keys:\n"
        "{\n"
        '  "consistency_flags": [ {"issue": str, "why_it_matters": str, "how_to_audit": str} ],\n'
        '  "narrative_summary_cn": str,\n'
        '  "qa_plan": [str]\n'
        "}\n\n"
        "Aggregated tables:\n"
        f"{summary_blob}"
    )

    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.2,
    )
    content = resp.choices[0].message.content.strip()
    content = re.sub(r"^```json\s*", "", content)
    content = re.sub(r"\s*```$", "", content)
    data = json.loads(content)

    out_path = os.path.join(out_dir, "final_ai_synthesis.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return data


# -----------------------------
# Excel Writer
# -----------------------------
def add_sheet_from_df(wb: Workbook, name: str, df: pd.DataFrame,
                      percent_cols: Optional[List[str]] = None,
                      freeze: str = "A3",
                      description: str = ""):
    """
    [Modified] 将 DataFrame 写入 Excel Sheet。
    改进点：
    1. 在第一行写入中文统计说明 (description)。
    2. 表头下移至第二行，数据从第三行开始。
    """
    ws = wb.create_sheet(title=name)

    # 1. 写入说明行 (Row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) if not df.empty else 1)
    cell_desc = ws.cell(row=1, column=1, value=description)
    cell_desc.font = Font(bold=True, italic=True, color="333333")
    cell_desc.fill = PatternFill("solid", fgColor="E0E0E0")
    cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30  # 增加高度以便显示说明

    # 2. 写入表头和数据 (从 Row 2 开始)
    # openpyxl 的 append 默认从第一行开始，我们需要手动处理
    rows = list(dataframe_to_rows(df, index=False, header=True))

    if not rows:
        return

    # 写入表头 (Row 2)
    header_row = rows[0]
    for col_idx, value in enumerate(header_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入数据 (Row 3+)
    for r_idx, row_data in enumerate(rows[1:], start=3):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 3. 设置冻结
    ws.freeze_panes = freeze

    # 4. 设置列宽和百分比格式
    for col_i, col_name in enumerate(df.columns, start=1):
        # 估算宽度
        sample = df[col_name].astype(str).head(50).tolist()
        max_len = max([len(str(col_name))] + [len(x) for x in sample]) if sample else len(str(col_name))
        ws.column_dimensions[get_column_letter(col_i)].width = min(max(10, max_len * 1.2), 60)

        # 百分比格式
        if percent_cols and col_name in percent_cols:
            col_letter = get_column_letter(col_i)
            # Apply to data rows only
            for r in range(3, ws.max_row + 1):
                ws.cell(row=r, column=col_i).number_format = "0.00%"


def write_final_excel(
        out_path: str,
        chunks: List[Chunk],
        records_df: pd.DataFrame,
        unparsed_df: pd.DataFrame,
        tables: Dict[str, pd.DataFrame],
        final_ai: Optional[Dict[str, Any]] = None
):
    print("正在写入 Excel (增强版)...")

    # 复制并清洗数据，防止 JSON 对象报错
    export_records = records_df.copy()
    export_unparsed = unparsed_df.copy()

    def serialize_complex_columns(df: pd.DataFrame):
        if df.empty: return df
        for col in df.columns:
            # 简单判断是否包含 list 或 dict
            if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
                df[col] = df[col].apply(
                    lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, (list, dict)) else x)
        return df

    export_records = serialize_complex_columns(export_records)
    export_unparsed = serialize_complex_columns(export_unparsed)

    wb = Workbook()
    wb.remove(wb.active)

    # --- 0. README ---
    ws = wb.create_sheet("README")
    lines = [
        "统计分析报告：系统功能语言学及物性分析",
        "------------------------------------------------",
        f"数据源：额尔古纳河右岸.pdf (正文)",
        f"统计范围：包含主句及嵌入小句 (Embedded Clauses)",
        "------------------------------------------------",
        "Sheet 说明：",
        "1. Summary_ProcessTypes: 全书及物性过程类型总体分布 (含嵌入小句)",
        "2. Summary_NatureSubcategory: 大自然主动参与者的内部构成 (动物/植物/环境)",
        "3. Summary_NatureVsHuman: 大自然 vs 人类 vs 人造物 的能动性对比",
        "4. NatureActive_RoleDist: 大自然作为不同语义角色 (Actor, Sayer等) 的分布",
        "5. NatureActive_ProcessDist: 大自然作为主动方参与的不同过程类型分布",
        "6. TopN_Animals_Detailed: 各类动物作为主动参与者的详细统计",
        "7. Annotations_All: 所有标注明细数据",
        "------------------------------------------------",
        "统计规则说明：",
        "- 动物部位 (如鹿角) 归并入 动物 (Animal)",
        "- 存在过程 (Existential) 不计入主动参与者 (Active Counted = 0)",
        "- 嵌入小句 (Embedded) 已计入过程类型总量统计",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    ws["A1"].font = Font(bold=True, size=14)

    # --- 1. 明细数据 ---
    # 分页写入防止超出限制
    chunk_df = pd.DataFrame([asdict(c) for c in chunks if hasattr(c, 'chunk_id')])  # 简单转换
    if "text_with_markers" in chunk_df.columns:
        chunk_df = chunk_df.drop(columns=["text_with_markers"])

    add_sheet_from_df(wb, "Index_Chunks", chunk_df, description="文本分块索引信息")

    if len(export_records) <= EXCEL_MAX_ROWS:
        add_sheet_from_df(wb, "Annotations_All", export_records,
                          description="[核心数据] 逐句标注明细，包含主句与嵌入小句")
    else:
        # Split logic omitted for brevity, same as original
        add_sheet_from_df(wb, "Annotations_All", export_records.head(EXCEL_MAX_ROWS),
                          description="[核心数据] 逐句标注明细 (Part 1)")

    if not export_unparsed.empty:
        add_sheet_from_df(wb, "Unparsed", export_unparsed, description="未成功解析的片段 (不计入统计)")

    # --- 2. 统计表 (带说明) ---

    # 定义每个表的中文说明映射
    table_descriptions = {
        "Summary_ProcessTypes": "【全书概览】所有及物性过程类型的数量与比例（分母=全书识别出的过程总数，含嵌入过程）。反映小说整体的叙事动态性。",
        "Summary_NatureSubcategory": "【大自然内部结构】在大自然作为主动参与者的事件中，动物、植物、自然环境各自的占比。",
        "Summary_NatureVsHuman": "【能动性对比】全书中“人类”、“大自然”、“人造物”作为主动参与者的数量对比。",
        "NatureActive_RoleDist": "【自然的角色】大自然在事件中充当的语义角色分布。",
        "NatureActive_ProcessDist": "【自然的行动】大自然作为主动方时，其参与的过程类型分布。",
        "Top10_SpeciesActive": "【活跃物种TOP10】全书中作为主动参与者出现频率最高的具体物种。",
        "TopN_Animals_Detailed": "【动物群像】各类动物作为主动参与者的详细统计。",
        # ========== 新增表格说明 ==========
        "Syntax_Ba_Bei": "【特种句式与受事赋权】统计小说中大自然作为物质过程动作者(Actor)时的“把字句”使用比例，以及大自然作为物质过程目标(Goal)时的“被字句”被动赋权比例。",
        "Syntax_Copula_Relational": "【关系与属性建构】统计大自然作为关系过程的载体(Carrier)时，系词('是', '像', '似', '无')的使用分布，这反映了拟人化、比喻或直接描写的叙事偏好。",
        "Circumstances_Dist": "【时空与方式环境】全文环境成分（时间、地点、方式）的数量分布对比，以及大自然占据主导能动性(Active)时的环境成分偏好。"
    }

    # 定义百分比列，自动进行 Excel 格式化
    percent_map = {
        "Summary_ProcessTypes": ["pct"],
        "Summary_NatureSubcategory": ["pct_of_nature"],
        "Summary_NatureVsHuman": ["pct_of_active"],
        "NatureActive_RoleDist": ["pct_of_nature_active"],
        "NatureActive_ProcessDist": ["pct_of_nature_active"],
        "Top10_SpeciesActive": ["pct_of_flora_fauna_active"],
        "TopN_Animals_Detailed": ["pct_of_all_animals"],
        # ========== 新增百分比列 ==========
        "Syntax_Ba_Bei": ["Ratio"],
        "Syntax_Copula_Relational": ["Ratio"],
        "Circumstances_Dist": ["Total_Ratio", "Nature_Active_Ratio"]
    }

    for name, df in tables.items():
        desc = table_descriptions.get(name, "统计数据表")
        pct_cols = percent_map.get(name, [])
        add_sheet_from_df(wb, name, df, percent_cols=pct_cols, description=desc)

    # --- 3. AI Synthesis ---
    if final_ai:
        ws2 = wb.create_sheet("AI_Synthesis_Report")
        ws2.column_dimensions["A"].width = 120
        ws2["A1"] = "AI 定性分析报告 (基于统计数据)"
        ws2["A1"].font = Font(bold=True, size=14)

        row = 3
        ws2[f"A{row}"] = "【叙事综述】"
        ws2[f"A{row}"].font = Font(bold=True)
        ws2[f"A{row + 1}"] = final_ai.get("narrative_summary_cn", "")

        row += 4
        ws2[f"A{row}"] = "【一致性检查与建议】"
        ws2[f"A{row}"].font = Font(bold=True)
        for item in final_ai.get("consistency_flags", []):
            row += 1
            ws2[f"A{row}"] = f"• 问题点: {item.get('issue', '')} (建议核查: {item.get('how_to_audit', '')})"

    wb.save(out_path)


# -----------------------------
# MAIN
# -----------------------------
def main():
    if not API_KEY:
        raise ValueError("API_KEY 为空。请在脚本顶部填入你的 key。")

    ensure_dir(OUT_DIR)
    cache_dir = os.path.join(OUT_DIR, "cache_chunks")
    ensure_dir(cache_dir)

    print("[1] Extract PDF text...")
    full_text = extract_pdf_text(PDF_PATH)

    print("[2] Slice narrative scope (skip author bio)...")
    scope_compact = slice_narrative_scope(full_text, START_SENT, END_SENT)

    print("[3] Sentence split...")
    sentences = sentence_split(scope_compact)
    # Safety: ensure first/last anchor presence
    if not sentences or (START_SENT.replace(" ", "") not in sentences[0].replace(" ", "")):
        print("WARN: 第一条句子未包含起点锚点，可能是PDF抽取差异，但仍继续。")
    if END_SENT.replace(" ", "") not in sentences[-1].replace(" ", ""):
        print("WARN: 最后一条句子未包含终点锚点，可能是PDF抽取差异，但仍继续。")

    print(f"Total sentences: {len(sentences)}")

    print("[4] Build chunks (paragraph-preferred, sentence-safe)...")
    chunks = build_chunks(sentences)
    print(f"Total chunks: {len(chunks)}")
    chunk_map = {c.chunk_id: c for c in chunks}

    print("[5] Call AI per chunk (with cache/resume)...")
    # client = make_client()
    check_ai_connectivity_or_exit()

    payloads = []
    failed_chunks = []
    for ch in chunks:
        try:
            payload = call_ai_for_chunk(ch, cache_dir)
            payloads.append(payload)
            print(
                f"  chunk {ch.chunk_id}: ok, records={len(payload.get('records', []))}, unparsed={len(payload.get('unparsed', []))}")
        except Exception as e:
            failed_chunks.append({"chunk_id": ch.chunk_id, "error": str(e)})
            print(f"  chunk {ch.chunk_id}: FAILED: {e}")

    # Save failures
    if failed_chunks:
        fail_path = os.path.join(OUT_DIR, "failed_chunks.json")
        with open(fail_path, "w", encoding="utf-8") as f:
            json.dump(failed_chunks, f, ensure_ascii=False, indent=2)
        print(f"WARN: some chunks failed. See: {fail_path}")

    if not payloads:
        raise RuntimeError("No successful chunk payloads; abort.")

    print("[6] Flatten records + compute tables...")
    records_df, unparsed_df = flatten_records(payloads, chunk_map)

    # Minimal normalization: ensure required columns exist
    for col in [
        "sent_id", "clause_id", "clause_text", "process_type", "process_subtype", "process_lexeme",
        "active_participant", "participant_category", "nature_subcategory", "canonical_flora_fauna",
        "role", "circumstances", "embedded_processes", "active_counted", "notes", "chunk_id"
    ]:
        if col not in records_df.columns:
            records_df[col] = None

    # Drop unparsed from stats by design (they are in unparsed_df)
    tables = compute_required_tables(records_df)

    print("[7] Final AI synthesis (optional but requested)...")
    final_ai = None
    # try:
    #     final_ai = call_ai_for_final_synthesis(client, tables, OUT_DIR)
    #     print("  final synthesis: ok")
    # except Exception as e:
    #     print(f"WARN: final synthesis failed: {e}")

    print("[8] Write Excel...")
    out_xlsx = os.path.join(OUT_DIR, "Ergun_FullNovel_SFL_Transitivity_v3.xlsx")
    write_final_excel(out_xlsx, chunks, records_df, unparsed_df, tables, final_ai=final_ai)

    print("DONE.")
    print(f"Excel saved to: {out_xlsx}")


if __name__ == "__main__":
    main()
