# -*- coding: utf-8 -*-
"""
Ergun River Narrative - SFL Transitivity Pipeline
- Extract narrative text from PDF (skip author bio via anchors)
- Chunk by paragraphs (fallback to sentences), ~3000-5000 Chinese chars
- Call OpenAI-compatible Chat Completions API (multiple independent tasks) with a stable coding manual
- Validate/repair JSON, retry failures, resume from cache
- Aggregate records, compute required stats, ask AI for final consistency + narrative synthesis
- Export final Excel to local disk

Dependencies:
  pip install pymupdf pandas openpyxl
"""

import os
import re
import json
import sys
import time
import math
import hashlib
from dataclasses import dataclass, asdict
from typing import List, Dict, Any, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ====== OpenAI SDK (OpenAI-compatible base_url) ======
from openai import OpenAI

# -----------------------------
# CONFIG
# -----------------------------
script_dir = os.path.dirname(os.path.abspath(__file__))
PDF_PATH = input("请输入PDF文件路径：").strip() or script_dir.strip('/') + "/doc/额尔古纳河右岸.pdf"
if not os.path.isfile(PDF_PATH):
    raise FileNotFoundError(f"PDF文件未找到：{PDF_PATH}")
OUT_DIR = os.path.dirname(PDF_PATH)   # 输出目录

# BASE_URL = "https://xiaoai.plus/v1/"
# BASE_URL = "http://127.0.0.1:23333/v1/"
# BASE_URL = "https://api.mttieeo.com/v1/"
# BASE_URL = "https://api.vectorengine.ai"
# BASE_URL = "https://ai.ttk.homes/v1"
BASE_URL = "https://api.deepseek.com"
API_KEY = input("请输入OpenAI API Key：").strip()  # 输入API Key
# API_KEY = "cs-sk-f0fd0228-4540-470f-8237-789684ac5f7e"
# MODEL = input("输入模型ID（如 gpt-4-turbo-0613）：").strip() or "gpt-4o"
# MODEL = "[满血]gemini-3.0-pro-preview"
# MODEL = "gemini-2.5-pro-thinking"
# MODEL = "gpt-5.1-thinking-all"
# MODEL = "gemini-3.1-pro-preview-cli"
# MODEL = "gemini-3-pro-preview-cli"
MODEL = "deepseek-reasoner"

# Narrative anchors
START_SENT = "我是雨和雪的老熟人了，我有九十岁了。"
END_SENT = "我落泪了，因为我已分不清天上人间了。"

# Chunk sizing (chars, after whitespace normalization; Chinese chars ~ tokens close to 1:1)
CHUNK_MIN_CHARS = 500
CHUNK_TARGET_CHARS = 1000
CHUNK_MAX_CHARS = 1500

# Retry policy
MAX_RETRIES = 4
RETRY_BACKOFF_SEC = 2.0

# Cache policy (resume)
USE_CACHE = True

# Excel sheet limits
EXCEL_MAX_ROWS = 1_000_000  # xlsx limit ~1,048,576

# -----------------------------
# CODING MANUAL / PROMPT
# -----------------------------
PROMPT_STAGE_1 = r"""
你是一个“系统功能语言学（SFL）”标注员。你需要对中文小说文本进行小句级的及物性标注。
【核心任务】提取过程、参与者分类，并标记其中包含的“地点环境成分”文本。

【强制输出格式】
使用纯文本块，每个小句用 [RECORD_START] 和 [RECORD_END] 包裹。
切勿输出真正的 JSON！不要有引号、逗号或注释。严格遵守以下字段：

[RECORD_START]
clause_text: (原文句子)
process_type: Material|Behavioural|Mental|Verbal|Relational|Existential
participant_category: Human|HumanBodyPart|Artifact|Nature|无
nature_subcategory: Animal|Plant|Environment|NatureOther|无
canonical_flora_fauna: (如：驯鹿、松树、额尔古纳河。若无填“无”)
role: Actor|Goal|Behaver|Senser|Sayer|Carrier|Existent|无
is_ba_sentence: False
is_bei_sentence: False
copula_type: (关系过程的系词，如：是、像、无)
active_counted: (1或0)
basic_places: (提取句中的地点词，如：在森林中, 神鼓上。用逗号分隔，无则填“无”)
[RECORD_END]
"""

PROMPT_STAGE_2 = r"""
你是一个空间语义与句法分析专家。
我将给你一组从小说中提取的句子和对应的【地点成分】。请你对该地点成分进行深度归类。

【深度分类规则】
1. 代词消解：如果是“那里/里面”，需根据上下文还原为原词。
2. 主类型：自然类 / 人类身体部位类 / 功能类(含建筑) / 地名类 / 方向类。
3. 树木信息：如果包含树/林，提取具体树种（如松树）和位置（如树梢）。
4. 句法结构：分析宏观语法（如：介词+处所词+方位词）和微观处所词语法（如：单一名词、名+的+名）。

【强制输出格式】
对于每个任务，严格使用以下纯文本格式：

[SPATIAL_START]
clause_text: (保持与原句一致)
circ_text: (保持与提取的地点一致)
resolved_target: (消解后的具体地点)
main_type: 自然类
sub_type: 植物 (或无)
center_word: 森林
markers: 在,中 (介词或方位词，逗号分隔，或填“无”)
tree_species: 桦树 (或无)
tree_position: 树下 (或无)
overall_syntax: 介词+处所词+方位词
noun_syntax: 单一名词
[SPATIAL_END]
"""


# -----------------------------
# Utilities
# -----------------------------
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def normalize_text_keep_para(text: str) -> str:
    """
    Keep paragraph boundaries, but normalize intra-line wraps and whitespace.
    Strategy:
      - Convert fullwidth spaces to normal removal
      - Keep blank lines as paragraph breaks
      - Remove spaces/tabs; merge wrapped lines within a paragraph
    """
    text = text.replace("\u3000", "")
    # unify newlines
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # remove trailing spaces per line
    lines = [re.sub(r"[ \t]+", "", ln) for ln in text.split("\n")]
    # rebuild paragraphs: blank line separates paragraphs
    paras = []
    buf = []
    for ln in lines:
        if ln == "":
            if buf:
                paras.append("".join(buf))
                buf = []
        else:
            buf.append(ln)
    if buf:
        paras.append("".join(buf))
    # remove empty paras
    paras = [p for p in paras if p.strip()]
    return "\n\n".join(paras)


def extract_pdf_text(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    pages = [doc.load_page(i).get_text("text") for i in range(doc.page_count)]
    return "\n".join(pages)


def slice_narrative_scope(full_text: str, start_sent: str, end_sent: str) -> str:
    """
    Use anchors to slice narrative正文，避开作者简介。
    End anchor may have whitespace breaks; we use regex with loose spaces removal.
    """
    # For robust matching, remove spaces/newlines but keep punctuation for anchor-finding
    compact = re.sub(r"[ \t\r\n\u3000]+", "", full_text)

    start_idx = compact.find(re.sub(r"[ \t\r\n\u3000]+", "", start_sent))
    if start_idx < 0:
        raise ValueError("未找到正文起点锚点（START_SENT）。请核对锚点句是否与PDF文本一致。")

    # For end, build tolerant regex on compact text
    end_compact = re.sub(r"[ \t\r\n\u3000]+", "", end_sent)
    end_idx = compact.find(end_compact)
    if end_idx < 0:
        # fallback: regex around key phrase “我落泪了...天上人间了。”
        m = re.search(r"我落泪了，?因为我已.*?分不清.*?天上人间了。", compact)
        if not m:
            raise ValueError("未找到正文终点锚点（END_SENT）。请核对终止句。")
        end_idx = m.end()
    else:
        end_idx += len(end_compact)

    scope_compact = compact[start_idx:end_idx]

    # Now we need a paragraph-preserving scope. Best effort:
    # We locate start/end in the original text by searching start/end in compact, then map back is hard.
    # Practical solution: extract text again paragraph-preserving, then locate using compact find within compact version and just use compact scope.
    # We will re-inject paragraph boundaries later by sentence markers, so compact scope is acceptable for analysis.
    return scope_compact


def sentence_split(text: str) -> List[str]:
    """
    Split by sentence-ending punctuation. Keep punctuation at end.
    """
    sents = []
    buf = ""
    for ch in text:
        buf += ch
        if ch in "。！？；":
            t = buf.strip()
            if t:
                sents.append(t)
            buf = ""
    if buf.strip():
        sents.append(buf.strip())
    return sents


def paragraph_split_from_sentences(sentences: List[str]) -> List[str]:
    """
    Recreate paragraphs roughly: since compact text loses original paras,
    we group sentences into paragraphs by heuristics:
      - If a sentence ends with '。' keep grouping; start new paragraph at dialogue markers or scene breaks.
    For chunking, paragraph boundaries are "soft"; we can just group every N sentences.
    """
    # Heuristic: new paragraph if sentence contains "——" or starts with "“" and previous ended with "。"
    paras = []
    buf = []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            paras.append("".join(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf:
        paras.append("".join(buf))
    return [p for p in paras if p.strip()]


@dataclass
class Chunk:
    chunk_id: int
    start_sent_id: int
    end_sent_id: int
    char_count: int
    text_with_markers: str
    sha1: str


def build_chunks(sentences: List[str]) -> List[Chunk]:
    """
    Chunk by paragraphs (heuristic), fallback to sentence boundaries.
    Add sentence markers to stabilize AI output: 〔S000123〕
    """
    paras = paragraph_split_from_sentences(sentences)

    # Build a mapping: paragraph -> sentence id range by reconstructing from concatenation
    # We'll just chunk by sentences directly but try to cut at para boundaries by using paras list
    # and tracking sentence pointers.
    chunks: List[Chunk] = []
    sent_idx = 1
    chunk_id = 1
    # Build para sentence lists: distribute sentences into paras in same order
    # Since paras were built from sentences sequentially, we can reconstruct by re-splitting paras (not needed).
    # We'll instead chunk via sentences with target sizes and allow "soft" paragraph breaks at para ends.
    # Precompute para boundaries in sentence indices:
    para_boundaries = []
    acc = 0
    # We reconstruct by iterating paras and counting how many sentences were in each by greedy matching
    # because paras were built from sentences with join, we can store counts during paragraph_split, but we didn't.
    # We'll re-do paragraph_split with counts:
    para_sent_counts = []
    buf = []
    for s in sentences:
        if buf and (("——" in s) or s.startswith("“") or s.startswith("『") or s.startswith("《")):
            para_sent_counts.append(len(buf))
            buf = [s]
        else:
            buf.append(s)
    if buf:
        para_sent_counts.append(len(buf))

    # Build para end indices
    end = 0
    for c in para_sent_counts:
        end += c
        para_boundaries.append(end)  # sentence index (1-based) where a paragraph ends

    def is_para_end(sid: int) -> bool:
        return sid in para_boundaries

    cur_sents = []
    cur_chars = 0
    chunk_start_sid = 1

    def flush_chunk(end_sid: int):
        nonlocal chunk_id, chunk_start_sid, cur_sents, cur_chars
        if not cur_sents:
            return
        # build text with sentence markers
        lines = []
        sid = chunk_start_sid
        for s in cur_sents:
            lines.append(f"〔S{sid:06d}〕{s}")
            sid += 1
        text_with_markers = "\n".join(lines)
        ch = Chunk(
            chunk_id=chunk_id,
            start_sent_id=chunk_start_sid,
            end_sent_id=end_sid,
            char_count=cur_chars,
            text_with_markers=text_with_markers,
            sha1=sha1_text(text_with_markers),
        )
        chunks.append(ch)
        chunk_id += 1
        chunk_start_sid = end_sid + 1
        cur_sents = []
        cur_chars = 0

    for sid, s in enumerate(sentences, start=1):
        s_len = len(s)
        # if adding would exceed max, flush before adding (ensure non-empty)
        if cur_sents and (cur_chars + s_len > CHUNK_MAX_CHARS):
            flush_chunk(sid - 1)

        cur_sents.append(s)
        cur_chars += s_len

        # flush when reaching target AND at paragraph end (preferred)
        if cur_chars >= CHUNK_TARGET_CHARS and is_para_end(sid):
            flush_chunk(sid)
        # or if very large, flush anyway at sentence end
        elif cur_chars >= CHUNK_MAX_CHARS:
            flush_chunk(sid)

    # remainder
    if cur_sents:
        flush_chunk(len(sentences))

    return chunks


# -----------------------------
# AI Call + JSON Validation
# -----------------------------
def make_client() -> OpenAI:
    # 显著增加 timeout，标注任务建议至少 120s，如果是慢速模型建议 300s
    # 同时增加 httpx 限制，防止连接过快断开
    import httpx
    http_client = httpx.Client(
        limits=httpx.Limits(max_connections=5, max_keepalive_connections=2),
        timeout=httpx.Timeout(300.0, read=240.0, connect=20.0)
    )
    client = OpenAI(
        base_url=BASE_URL,
        api_key=API_KEY,
        http_client=http_client
    )
    return client


def extract_visible_text_from_choice_message(msg) -> str:
    """
    Compatible with normal models (message.content) and reasoning models
    (message.reasoning_content). Return the best-effort visible text.
    """
    content = getattr(msg, "content", None)
    if isinstance(content, str) and content.strip():
        return content.strip()

    # some reasoning models put output here
    rc = getattr(msg, "reasoning_content", None)
    if isinstance(rc, str) and rc.strip():
        return rc.strip()

    # fallback: stringify
    try:
        s = str(msg)
        return s.strip()
    except Exception:
        return ""


def check_ai_connectivity_or_exit():
    """
    使用原生 requests 检查连接
    """
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": "say ok"}],
        "max_tokens": 5
    }
    try:
        # 使用较短的 timeout 进行测试
        response = requests.post(f"{BASE_URL.rstrip('/')}/chat/completions",
                                 headers=headers, json=payload, timeout=20)
        response.raise_for_status()
        print("  [Connect] API 连接成功")
    except Exception as e:
        print(f"\n[ERROR] 接口访问失败: {e}")
        if response := getattr(e, 'response', None):
            print(f"状态码: {response.status_code}, 返回内容: {response.text}")
        raise SystemExit(1)


ALLOWED_PROCESS_TYPES = {"Material", "Behavioural", "Mental", "Verbal", "Relational", "Existential"}
ALLOWED_PARTICIPANT_CATS = {"Human", "HumanBodyPart", "Artifact", "Nature"}
ALLOWED_ROLES = {"Actor", "Goal", "Behaver", "Senser", "Sayer", "Carrier", "Existent"}


def basic_validate_payload(payload: Dict[str, Any]) -> Tuple[bool, List[str]]:
    errs = []
    if not isinstance(payload, dict):
        return False, ["payload_not_dict"]
    if "chunk_id" not in payload or not isinstance(payload["chunk_id"], int):
        errs.append("missing_or_bad_chunk_id")
    if "records" not in payload or not isinstance(payload["records"], list):
        errs.append("missing_or_bad_records")

    for i, r in enumerate(payload.get("records", [])):
        if not isinstance(r, dict):
            errs.append(f"record_{i}_not_dict")
            continue
        # 移除了过于严苛的必填字段检查，因为有些模型偶尔会漏掉新增的可选布尔字段，防止重试风暴
        pt = r.get("process_type")
        if pt and pt not in ALLOWED_PROCESS_TYPES:
            errs.append(f"record_{i}_bad_process_type:{pt}")
        pc = r.get("participant_category")
        if pc and pc not in ALLOWED_PARTICIPANT_CATS:
            errs.append(f"record_{i}_bad_participant_category:{pc}")
        role = r.get("role")
        if role and role not in ALLOWED_ROLES:
            errs.append(f"record_{i}_bad_role:{role}")

    return (len(errs) == 0), errs


def is_cloudflare_524_or_timeout(err: Exception) -> bool:
    """
    Best-effort detection for Cloudflare 524 / timeouts from OpenAI-compatible SDKs.
    """
    s = str(err).lower()
    status = getattr(err, "status_code", None)
    if status == 524:
        return True
    if "error code 524" in s or "cloudflare" in s and "524" in s:
        return True
    if "timeout" in s or "timed out" in s or "readtimeout" in s:
        return True
    return False


def split_marked_text_by_lines(marked_text: str, sub_max_chars: int = 1800) -> List[str]:
    """
    marked_text is multiple lines like: 〔S000123〕....
    Split into smaller parts without breaking a line.
    """
    lines = [ln for ln in marked_text.split("\n") if ln.strip()]
    parts = []
    buf = []
    cur = 0
    for ln in lines:
        ln_len = len(ln)
        if buf and (cur + ln_len > sub_max_chars):
            parts.append("\n".join(buf))
            buf = [ln]
            cur = ln_len
        else:
            buf.append(ln)
            cur += ln_len
    if buf:
        parts.append("\n".join(buf))
    return parts


def clean_ai_value(v: str) -> str:
    """强力清洗大模型可能偷偷加上去的 JSON 伪影 (逗号、引号、注释)"""
    v = re.sub(r'//.*$', '', v)  # 移除行尾的双斜杠注释
    v = v.strip()
    v = re.sub(r',$', '', v)  # 移除末尾逗号
    v = re.sub(r'^["\']|["\']$', '', v)  # 移除首尾引号
    return v.strip()


def parse_stage1_text(raw_text: str, chunk_id: int) -> Dict[str, Any]:
    records = []
    unparsed = []
    blocks = re.findall(r'\[RECORD_START\](.*?)\[RECORD_END\]', raw_text, re.DOTALL)

    if not blocks:
        # 强化 fallback 逻辑：过滤掉无意义的连接词（如“和”）
        raw_blocks = raw_text.split('[RECORD_START]')
        blocks = []
        for b in raw_blocks:
            b = b.strip()
            # 剥离可能残存的尾部标签
            b = re.sub(r'\[RECORD_END\].*', '', b, flags=re.DOTALL).strip()
            # 【核心修复】：只有包含冒号（中英文均可）的块，才被认为是有效的数据块
            if b and (":" in b or "：" in b):
                blocks.append(b)

    for block in blocks:
        record = {"chunk_id": chunk_id, "circumstances": []}
        places_raw = ""
        lines = block.strip().split('\n')
        # ... (内部逻辑保持原有不变)
        for line in lines:
            match = re.match(r'^"?([a-zA-Z_]+)"?\s*[:：]\s*(.*)$', line.strip())
            if match:
                k, v = match.groups()
                v = clean_ai_value(v)

                if v.lower() in ['null', 'none', '无', '-', '']:
                    v = None
                elif v.lower() in ['true', '是']:
                    v = True
                elif v.lower() in ['false', '否']:
                    v = False

                if k == "basic_places":
                    places_raw = v
                else:
                    record[k] = v

        if record.get("clause_text"):
            if places_raw and places_raw not in ["无", "None"]:
                place_list = [p.strip() for p in places_raw.replace('，', ',').split(',') if p.strip()]
                record["_pending_places"] = place_list
            records.append(record)
        else:
            unparsed.append({"raw_data": block[:50], "error_reason": "Missing clause_text"})

    # 【核心修复】：将 chunk_id 加回到根目录
    return {"chunk_id": chunk_id, "records": records, "unparsed": unparsed}


def parse_stage2_text(raw_text: str) -> Dict[str, Dict]:
    """解析第二阶段的空间详细信息，返回以 (clause_text, circ_text) 为键的字典"""
    spatial_map = {}
    blocks = re.findall(r'\[SPATIAL_START\](.*?)\[SPATIAL_END\]', raw_text, re.DOTALL)
    for block in blocks:
        sd = {}
        lines = block.strip().split('\n')
        for line in lines:
            match = re.match(r'^"?([a-zA-Z_]+)"?\s*[:：]\s*(.*)$', line.strip())
            if match:
                k, v = match.groups()
                v = clean_ai_value(v)
                if v.lower() in ['null', 'none', '无', '-', '']: v = None
                sd[k] = v

        clause_txt = sd.get("clause_text", "")
        circ_txt = sd.get("circ_text", "")
        if clause_txt and circ_txt:
            # 处理 markers
            raw_markers = sd.get("markers", "")
            markers_list = []
            if raw_markers and raw_markers != "无":
                markers_list = [m.strip() for m in raw_markers.replace('，', ',').split(",") if m.strip()]
            sd["markers"] = markers_list
            spatial_map[f"{clause_txt}_{circ_txt}"] = sd

    return spatial_map


def call_ai_for_chunk(chunk: Chunk, cache_dir: str) -> Dict[str, Any]:
    """
    两阶段流水线架构 (Two-Stage Pipeline):
    1. 提取核心 SFL 和地点列表
    2. 如果有地点，单独针对地点进行深入请求
    3. 合并返回，极大降低大模型幻觉和截断率
    """
    ensure_dir(cache_dir)
    cache_path = os.path.join(cache_dir, f"chunk_{chunk.chunk_id:03d}_{chunk.sha1}.json")

    if USE_CACHE and os.path.exists(cache_path):
        with open(cache_path, "r", encoding="utf-8") as f:
            print(f"  [chunk {chunk.chunk_id}] 读取缓存")
            return json.load(f)

    def _call_api_stream(system_prompt: str, user_prompt: str, task_name: str) -> str:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": 0,
            "stream": True
        }

        full_content = []
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                if attempt > 1:
                    print(f"\n    {task_name} 重试 {attempt}/{MAX_RETRIES}... ", end="", flush=True)

                response = requests.post(f"{BASE_URL.rstrip('/')}/chat/completions", headers=headers, json=payload,
                                         timeout=(15, 600), stream=True)
                response.raise_for_status()

                char_counter = 0
                for line in response.iter_lines():
                    if not line: continue
                    line_text = line.decode("utf-8")
                    if line_text.startswith("data: ") and "[DONE]" not in line_text:
                        try:
                            data_json = json.loads(line_text[6:])
                            if "error" in data_json: raise RuntimeError(f"API报错: {data_json.get('error')}")
                            delta = data_json["choices"][0].get("delta", {})

                            piece = delta.get("content") or delta.get("reasoning_content")
                            if piece:
                                full_content.append(piece)
                                char_counter += len(piece)
                                if char_counter > 30:
                                    sys.stdout.write("." if delta.get("content") else "*")
                                    sys.stdout.flush()
                                    char_counter = 0
                        except Exception:
                            pass
                print(" 完成")

                raw = "".join(full_content).strip()
                # 强力剔除思考标签
                clean = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
                return clean

            except Exception as e:
                print(f" [出错: {str(e)[:50]}]", end="", flush=True)
                time.sleep(RETRY_BACKOFF_SEC)
                full_content = []

        raise RuntimeError(f"{task_name} 重试耗尽")

    print(f"    处理 chunk {chunk.chunk_id} ", end="", flush=True)

    # ============ STAGE 1: 核心提取 ============
    s1_prompt = f"{PROMPT_STAGE_1}\n\n【分析文本】\n{chunk.text_with_markers}"
    s1_raw = _call_api_stream("严格遵循格式，禁止输出JSON符号。", s1_prompt, "Stage-1")
    parsed_stage1 = parse_stage1_text(s1_raw, chunk.chunk_id)

    # 提取需要深度分析的地名
    places_to_analyze = []
    for rec in parsed_stage1["records"]:
        if "_pending_places" in rec:
            for p in rec["_pending_places"]:
                places_to_analyze.append(f"- 原句：{rec.get('clause_text')}\n  地点：{p}")

    # ============ STAGE 2: 深度空间解析 (按需) ============
    spatial_map = {}
    if places_to_analyze:
        print(f"      -> 发现 {len(places_to_analyze)} 个地点成分，启动深度解析 ", end="", flush=True)
        places_str = "\n".join(places_to_analyze)
        s2_prompt = f"{PROMPT_STAGE_2}\n\n【需分析的句子与地点列表】\n{places_str}"
        s2_raw = _call_api_stream("专注空间句法分析。", s2_prompt, "Stage-2")
        spatial_map = parse_stage2_text(s2_raw)

    # ============ 组装合并 (Merge) ============
    for rec in parsed_stage1["records"]:
        pending = rec.pop("_pending_places", [])
        for p in pending:
            key = f"{rec.get('clause_text')}_{p}"
            sd = spatial_map.get(key)
            if sd:
                # 组装为后续 pandas 统计模块需要的格式
                rec["circumstances"].append({
                    "type": "Place",
                    "text": p,
                    "spatial_details": sd
                })
            else:
                # 兜底：如果 Stage 2 失败或遗漏，保留浅层提取
                rec["circumstances"].append({"type": "Place", "text": p, "spatial_details": None})

    # 缓存并返回
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(parsed_stage1, f, ensure_ascii=False, indent=2)

    return parsed_stage1



def repair_json_via_ai(client: OpenAI, bad_text: str, errs: List[str]) -> Dict[str, Any]:
    """
    Ask AI to output valid JSON only, repairing prior output.
    """
    system_msg = "You are a strict JSON repair tool. Output ONLY valid JSON."
    user_msg = (
        "The previous output is not valid per schema. "
        f"Errors: {errs[:20]}\n"
        "Rewrite it into a STRICTLY valid JSON matching the schema. "
        "Do not add extra commentary.\n\n"
        f"Previous output:\n{bad_text}"
    )
    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.0,
    )
    content = resp.choices[0].message.content.strip()
    content = re.sub(r"^```json\s*", "", content)
    content = re.sub(r"\s*```$", "", content)
    return json.loads(content)


# -----------------------------
# Aggregation & Stats
# -----------------------------
def flatten_records(all_payloads: List[Dict[str, Any]], chunk_map: Dict[int, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    将嵌套的 JSON 展平为 DataFrame。
    [健壮性增强版]：修复数据类型不匹配，并兼容老版本缓存中丢失的 chunk_id。
    """
    recs = []
    unparsed = []

    for payload in all_payloads:
        # 尝试从 payload 获取，如果没有，从内部记录恢复（完美兼容你的现有缓存）
        cid = payload.get("chunk_id", -1)
        if cid == -1 and payload.get("records"):
            cid = payload["records"][0].get("chunk_id", -1)

        ch = chunk_map.get(cid)

        # 1. 处理主句记录
        for r in payload.get("records", []):
            if not isinstance(r, dict):
                continue

            rec = dict(r)
            rec["chunk_id"] = cid
            rec["chunk_start_sent_id"] = ch.start_sent_id if ch else None
            rec["chunk_end_sent_id"] = ch.end_sent_id if ch else None
            rec["is_embedded"] = False

            # 【核心修复】：强制类型转换，将字符串 "1" 转换为整数 1
            try:
                rec["active_counted"] = int(rec.get("active_counted", 0))
            except (ValueError, TypeError):
                rec["active_counted"] = 0

            recs.append(rec)

            # 2. 处理嵌入小句
            if "embedded_processes" in r and isinstance(r["embedded_processes"], list):
                for emb in r["embedded_processes"]:
                    if not isinstance(emb, dict): continue
                    emb_rec = {
                        "chunk_id": cid,
                        "sent_id": rec.get("sent_id"),
                        "clause_id": f"{rec.get('clause_id', '')}_emb",
                        "clause_text": emb.get("text", "") or emb.get("clause_text", ""),
                        "process_type": emb.get("process_type"),
                        "active_counted": 0, # 嵌入小句不计入 active
                        "is_embedded": True
                    }
                    recs.append(emb_rec)

        # 3. 处理解析失败的片段
        for u in payload.get("unparsed", []):
            if isinstance(u, dict):
                item = dict(u)
            else:
                item = {"raw_data": str(u), "error_reason": "Format Error"}
            item["chunk_id"] = cid
            unparsed.append(item)

    records_df = pd.DataFrame(recs)
    unparsed_df = pd.DataFrame(unparsed)
    return records_df, unparsed_df


def safe_pct(n: int, d: int) -> float:
    return (n / d) if d else 0.0


def compute_spatial_tables(records_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """处理并计算所有的地点环境成分统计需求 (需求 1 - 9)"""
    spatial_list = []

    # 提取所有地点成分，并与主句的 Participant 绑定 (为需求 9 准备)
    for idx, row in records_df.iterrows():
        circs = row.get("circumstances", [])
        participant_cat = row.get("participant_category", "Unknown")

        if isinstance(circs, list):
            for c in circs:
                if isinstance(c, dict) and c.get("spatial_details"):
                    sd = c["spatial_details"]
                    spatial_list.append({
                        "clause_text": row.get("clause_text"),
                        "text": c.get("text", ""),
                        "resolved_target": sd.get("resolved_target", ""),
                        "main_type": sd.get("main_type", "未分类"),
                        "sub_type": sd.get("sub_type", "未分类"),
                        "center_word": sd.get("center_word", ""),
                        "markers": sd.get("markers", []),
                        "tree_species": sd.get("tree_species"),
                        "tree_position": sd.get("tree_position"),
                        "overall_syntax": sd.get("overall_syntax", "未识别"),
                        "noun_syntax": sd.get("noun_syntax", "未识别"),
                        "parent_participant": participant_cat
                    })

    spatial_df = pd.DataFrame(spatial_list)
    if spatial_df.empty:
        return {}

    tables = {}

    # 需求 1 & 2: 数量、类型、子类型、比例
    total_spatial = len(spatial_df)
    type_counts = spatial_df.groupby(["main_type", "sub_type"]).size().reset_index(name="Count")
    type_counts["Ratio"] = type_counts["Count"] / total_spatial
    type_counts = type_counts.sort_values(by=["main_type", "Count"], ascending=[True, False])
    tables["Spatial_Types_Dist"] = type_counts

    # 准备：筛选自然类
    nature_df = spatial_df[spatial_df["main_type"] == "自然类"].copy()
    total_nature = len(nature_df)

    # 需求 3: 自然类中 河、山、树、地、林、水的数量和比例
    keywords = ["河", "山", "树", "地", "林", "水"]
    kw_stats = []
    if total_nature > 0:
        for kw in keywords:
            # 匹配 text 或消解后的 target
            count = nature_df["resolved_target"].apply(lambda x: kw in str(x)).sum()
            kw_stats.append({"Keyword": kw, "Count": count, "Ratio": count / total_nature if total_nature else 0})
    tables["Spatial_Nature_Keywords"] = pd.DataFrame(kw_stats)

    # 需求 4: 自然类地点中排名前10的地点
    if total_nature > 0:
        top10_places = nature_df["resolved_target"].value_counts().head(10).reset_index()
        top10_places.columns = ["Resolved_Target", "Count"]
        top10_places["Ratio"] = top10_places["Count"] / total_nature
        tables["Spatial_Nature_Top10"] = top10_places

    # 需求 5: 方位词和介词的统计
    target_markers = ["前", "后", "上", "下", "左", "右", "东", "南", "西", "北", "中", "里", "边", "旁", "外", "在",
                      "朝", "向", "到"]
    marker_counts = {m: 0 for m in target_markers}
    for markers in spatial_df["markers"].dropna():
        if isinstance(markers, list):
            for m in markers:
                for tm in target_markers:
                    if tm in str(m):
                        marker_counts[tm] += 1
    marker_df = pd.DataFrame(list(marker_counts.items()), columns=["Marker", "Count"])
    marker_df["Ratio"] = marker_df["Count"] / total_spatial if total_spatial else 0
    tables["Spatial_Markers_Dist"] = marker_df.sort_values(by="Count", ascending=False)

    # 需求 6: 树/林相关细节
    tree_df = nature_df[nature_df["resolved_target"].astype(str).str.contains("树|林", regex=True)]
    total_trees = len(tree_df)
    if total_trees > 0:
        # 树种分布
        species_df = tree_df["tree_species"].dropna().value_counts().reset_index()
        species_df.columns = ["Tree_Species", "Species_Count"]
        species_df["Species_Ratio"] = species_df["Species_Count"] / total_trees
        # 树的位置分布
        position_df = tree_df["tree_position"].dropna().value_counts().reset_index()
        position_df.columns = ["Tree_Position", "Position_Count"]
        position_df["Position_Ratio"] = position_df["Position_Count"] / total_trees
        tables["Spatial_Tree_Details"] = pd.concat([species_df, position_df], axis=1)  # 并排展示方便查看

    # 需求 7: 自然类整体语法结构 (overall_syntax)
    if total_nature > 0:
        overall_syn = nature_df["overall_syntax"].value_counts().reset_index()
        overall_syn.columns = ["Overall_Syntax", "Count"]
        overall_syn["Ratio"] = overall_syn["Count"] / total_nature
        tables["Spatial_Nature_OverallSyntax"] = overall_syn

    # 需求 8: 处所词语法结构 (noun_syntax)
    if total_spatial > 0:
        noun_syn = spatial_df["noun_syntax"].value_counts().reset_index()
        noun_syn.columns = ["Noun_Syntax", "Count"]
        noun_syn["Ratio"] = noun_syn["Count"] / total_spatial
        tables["Spatial_All_NounSyntax"] = noun_syn

    # 需求 9: 自然类地点同现的参与者分布 (人类 vs 自然类)
    if total_nature > 0:
        participant_dist = nature_df["parent_participant"].value_counts().reset_index()
        participant_dist.columns = ["Participant_Category", "Count"]
        participant_dist["Ratio"] = participant_dist["Count"] / total_nature
        # 仅保留 Human 和 Nature
        participant_dist = participant_dist[participant_dist["Participant_Category"].isin(["Human", "Nature"])]
        tables["Spatial_CoOccur_Participants"] = participant_dist

    return tables


def compute_required_tables(records_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    if records_df.empty:
        raise ValueError("records_df is empty; no parsed records to summarize.")

        # ================= 数据类型与格式安全锁 =================
        # 确保参与者类型前后没有空格导致匹配失败
    if "participant_category" in records_df.columns:
        records_df["participant_category"] = records_df["participant_category"].astype(str).str.strip()

        # 为防止任何漏网之鱼，在 Pandas 层面再次确保 active_counted 绝对是 int
    if "active_counted" in records_df.columns:
        records_df["active_counted"] = pd.to_numeric(records_df["active_counted"], errors='coerce').fillna(0).astype(
            int)
    # ==========================================================

    # 补全列名防止报错（包括新增的句法特征列）
    expected_cols = [
        "process_type", "participant_category", "nature_subcategory",
        "canonical_flora_fauna", "role", "active_counted", "is_embedded",
        "is_ba_sentence", "is_bei_sentence", "copula_type", "circumstances"
    ]
    for col in expected_cols:
        if col not in records_df.columns:
            # 布尔列默认 False，其他默认 None
            if col in ["is_ba_sentence", "is_bei_sentence"]:
                records_df[col] = False
            else:
                records_df[col] = None

    # ---- 基础表：前置原有逻辑 (已精简提取) ----
    proc = records_df[records_df["process_type"].notna()].groupby("process_type").size().reset_index(
        name="n").sort_values("n", ascending=False)
    proc["pct"] = proc["n"] / proc["n"].sum()

    active_df = records_df[records_df["active_counted"] == 1].copy()
    n_active = len(active_df)
    nature_all = active_df[active_df["participant_category"] == "Nature"].copy()
    n_nature = len(nature_all)

    nature_sub_sum = nature_all.groupby("nature_subcategory").size().reset_index(name="n").sort_values("n",
                                                                                                       ascending=False) if n_nature > 0 else pd.DataFrame(
        columns=["nature_subcategory", "n", "pct_of_nature"])
    if not nature_sub_sum.empty: nature_sub_sum["pct_of_nature"] = nature_sub_sum["n"] / n_nature

    cat_summary = pd.DataFrame([
        {"Category": "Nature", "n": len(active_df[active_df["participant_category"] == "Nature"])},
        {"Category": "Human (incl. BodyPart)",
         "n": len(active_df[active_df["participant_category"].isin(["Human", "HumanBodyPart"])])},
        {"Category": "Artifact", "n": len(active_df[active_df["participant_category"] == "Artifact"])}
    ])
    cat_summary["pct_of_active"] = cat_summary["n"] / n_active if n_active else 0.0

    role_sum = nature_all.groupby("role").size().reset_index(name="n").sort_values("n", ascending=False)
    role_sum["pct_of_nature_active"] = role_sum["n"] / n_nature if n_nature else 0.0

    nature_proc = nature_all.groupby("process_type").size().reset_index(name="n").sort_values("n", ascending=False)
    nature_proc["pct_of_nature_active"] = nature_proc["n"] / n_nature if n_nature else 0.0

    ff = nature_all[nature_all["canonical_flora_fauna"].notna()].copy()
    top10 = ff.groupby("canonical_flora_fauna").size().reset_index(name="n").sort_values("n", ascending=False).head(10)
    top10["pct_of_flora_fauna_active"] = top10["n"] / len(ff) if len(ff) else 0.0

    # (省略TopN详细代码以节省空间，直接使用原逻辑)
    TOPN_ANIMALS = 15
    animal_ff = ff[ff["nature_subcategory"] == "Animal"].copy()
    if not animal_ff.empty:
        animal_totals = animal_ff.groupby("canonical_flora_fauna").size().reset_index(name="Total").sort_values("Total",
                                                                                                                ascending=False)
        animal_totals["pct_of_all_animals"] = animal_totals["Total"] / animal_totals["Total"].sum()
        top_animals = animal_totals.head(TOPN_ANIMALS)["canonical_flora_fauna"].tolist()
        animal_top_df = animal_ff[animal_ff["canonical_flora_fauna"].isin(top_animals)]
        animal_role_counts = animal_top_df.pivot_table(index="canonical_flora_fauna", columns="role", values="sent_id",
                                                       aggfunc="count", fill_value=0).reset_index()
        for rc in ["Actor", "Behaver", "Senser", "Sayer", "Carrier"]:
            if rc not in animal_role_counts.columns: animal_role_counts[rc] = 0
        TopN_Animals_ByRole = animal_totals[animal_totals["canonical_flora_fauna"].isin(top_animals)].merge(
            animal_role_counts, on="canonical_flora_fauna", how="left").sort_values("Total", ascending=False)
    else:
        TopN_Animals_ByRole = pd.DataFrame(
            columns=["canonical_flora_fauna", "Total", "pct_of_all_animals", "Actor", "Behaver", "Senser", "Sayer",
                     "Carrier"])

    # ================= 新增分析模块 =================

    # ---- 新增 1 & 2: 物质过程中的把字句 (Actor) 和被字句 (Goal) ----
    nature_mat_actor = records_df[
        (records_df["process_type"] == "Material") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Actor")]
    ba_total = len(nature_mat_actor)
    ba_count = nature_mat_actor["is_ba_sentence"].sum() if ba_total > 0 else 0
    ba_stats = pd.DataFrame([{
        "Analysis_Type": "大自然作为物质过程动作者(Actor)",
        "Total_Instances": ba_total,
        "Target_Sentence_Type": "把字句/将字句",
        "Count": ba_count,
        "Ratio": ba_count / ba_total if ba_total > 0 else 0.0
    }])

    nature_mat_goal = records_df[
        (records_df["process_type"] == "Material") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Goal")]
    bei_total = len(nature_mat_goal)
    bei_count = nature_mat_goal["is_bei_sentence"].sum() if bei_total > 0 else 0
    bei_stats = pd.DataFrame([{
        "Analysis_Type": "大自然作为物质过程目标(Goal)",
        "Total_Instances": bei_total,
        "Target_Sentence_Type": "被动句式",
        "Count": bei_count,
        "Ratio": bei_count / bei_total if bei_total > 0 else 0.0
    }])
    syntax_ba_bei = pd.concat([ba_stats, bei_stats], ignore_index=True)

    # ---- 新增 3: 关系过程中的系词分布 ----
    nature_carrier = records_df[
        (records_df["process_type"] == "Relational") & (records_df["participant_category"] == "Nature") & (
                    records_df["role"] == "Carrier")]
    carrier_total = len(nature_carrier)
    if carrier_total > 0:
        # 将空值填充为"未识别"
        nature_carrier.loc[:, "copula_type"] = nature_carrier["copula_type"].fillna("未识别")
        copula_stats = nature_carrier["copula_type"].value_counts().reset_index()
        copula_stats.columns = ["Copula_Type (系词类型)", "Count"]
        copula_stats["Ratio"] = copula_stats["Count"] / carrier_total
    else:
        copula_stats = pd.DataFrame(columns=["Copula_Type (系词类型)", "Count", "Ratio"])

    # ---- 新增 4: 环境成分 (Circumstances) 分布 ----
    circ_list = []
    for idx, row in records_df.iterrows():
        circs = row.get("circumstances", [])
        is_nature = (row.get("participant_category") == "Nature")
        # 宽泛匹配，哪怕 Nature 不是 Active，只要它出现在句子中我们就打标
        is_nature_active = is_nature and row.get("active_counted") == 1

        if isinstance(circs, list):
            for c in circs:
                if isinstance(c, dict):
                    ctype = str(c.get("type", "")).lower()
                    csubtype = str(c.get("subtype", "")).lower()

                    # 归一化为三大类
                    norm_type = "Other"
                    if "time" in ctype or "time" in csubtype or "extent" in ctype:
                        norm_type = "Time (时间/时长)"
                    elif "place" in ctype or "place" in csubtype or "location" in ctype:
                        norm_type = "Place (地点/空间)"
                    elif "manner" in ctype:
                        norm_type = "Manner (方式)"

                    circ_list.append({
                        "Circumstance_Type": norm_type,
                        "Is_Nature_Active": is_nature_active
                    })

    circ_df = pd.DataFrame(circ_list)
    if not circ_df.empty:
        circ_all = circ_df["Circumstance_Type"].value_counts().reset_index()
        circ_all.columns = ["Circumstance_Type", "Total_Count"]
        circ_all["Total_Ratio"] = circ_all["Total_Count"] / circ_all["Total_Count"].sum()

        circ_nature = circ_df[circ_df["Is_Nature_Active"]]["Circumstance_Type"].value_counts().reset_index()
        circ_nature.columns = ["Circumstance_Type", "Nature_Active_Count"]

        circ_stats = pd.merge(circ_all, circ_nature, on="Circumstance_Type", how="left").fillna(0)
        total_nature_circs = circ_stats["Nature_Active_Count"].sum()
        circ_stats["Nature_Active_Ratio"] = circ_stats[
                                                "Nature_Active_Count"] / total_nature_circs if total_nature_circs > 0 else 0.0
    else:
        circ_stats = pd.DataFrame(
            columns=["Circumstance_Type", "Total_Count", "Total_Ratio", "Nature_Active_Count", "Nature_Active_Ratio"])

    base_tables = {
        "Summary_ProcessTypes": proc,
        "Summary_NatureSubcategory": nature_sub_sum,
        "Summary_NatureVsHuman": cat_summary,
        "NatureActive_RoleDist": role_sum,
        "NatureActive_ProcessDist": nature_proc,
        "Top10_SpeciesActive": top10,
        "TopN_Animals_Detailed": TopN_Animals_ByRole,
        "Syntax_Ba_Bei": syntax_ba_bei,
        "Syntax_Copula_Relational": copula_stats,
        "Circumstances_Dist": circ_stats
    }

    # 将新增的地点成分表格合并进来
    spatial_tables = compute_spatial_tables(records_df)
    base_tables.update(spatial_tables)

    return base_tables



# -----------------------------
# Final AI synthesis (optional but requested)
# -----------------------------
def call_ai_for_final_synthesis(client: OpenAI, tables: Dict[str, pd.DataFrame], out_dir: str) -> Dict[str, Any]:
    """
    Send summarized tables (not full records) back to AI for:
      - consistency check suggestions
      - narrative ecological linguistics insight
    Output JSON and save.
    """

    def df_to_markdown(df: pd.DataFrame, max_rows=50) -> str:
        return df.head(max_rows).to_markdown(index=False)

    payload_md = []
    for name, df in tables.items():
        payload_md.append(f"## {name}\n{df_to_markdown(df)}\n")
    summary_blob = "\n".join(payload_md)

    system_msg = "You are a senior SFL + ecolinguistics reviewer. Output ONLY JSON."
    user_msg = (
        "You will receive aggregated statistics from a clause-level SFL transitivity annotation of a Chinese eco-narrative.\n"
        "Tasks:\n"
        "1) Check internal consistency and flag likely coding inconsistencies to audit (e.g., 有 existential vs possessive; 看/看见; 被字句 active_counted).\n"
        "2) Provide a concise narrative synthesis (Chinese) of what the distributions imply for 'nature agency' in the text.\n"
        "3) Provide concrete next-step QA suggestions (what to spot-check).\n"
        "Return STRICT JSON with keys:\n"
        "{\n"
        '  "consistency_flags": [ {"issue": str, "why_it_matters": str, "how_to_audit": str} ],\n'
        '  "narrative_summary_cn": str,\n'
        '  "qa_plan": [str]\n'
        "}\n\n"
        "Aggregated tables:\n"
        f"{summary_blob}"
    )

    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.2,
    )
    content = resp.choices[0].message.content.strip()
    content = re.sub(r"^```json\s*", "", content)
    content = re.sub(r"\s*```$", "", content)
    data = json.loads(content)

    out_path = os.path.join(out_dir, "final_ai_synthesis.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return data


# -----------------------------
# Excel Writer
# -----------------------------
def add_sheet_from_df(wb: Workbook, name: str, df: pd.DataFrame,
                      percent_cols: Optional[List[str]] = None,
                      freeze: str = "A3",
                      description: str = ""):
    """
    [Modified] 将 DataFrame 写入 Excel Sheet。
    改进点：
    1. 在第一行写入中文统计说明 (description)。
    2. 表头下移至第二行，数据从第三行开始。
    """
    ws = wb.create_sheet(title=name)

    # 1. 写入说明行 (Row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) if not df.empty else 1)
    cell_desc = ws.cell(row=1, column=1, value=description)
    cell_desc.font = Font(bold=True, italic=True, color="333333")
    cell_desc.fill = PatternFill("solid", fgColor="E0E0E0")
    cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30  # 增加高度以便显示说明

    # 2. 写入表头和数据 (从 Row 2 开始)
    # openpyxl 的 append 默认从第一行开始，我们需要手动处理
    rows = list(dataframe_to_rows(df, index=False, header=True))

    if not rows:
        return

    # 写入表头 (Row 2)
    header_row = rows[0]
    for col_idx, value in enumerate(header_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入数据 (Row 3+)
    for r_idx, row_data in enumerate(rows[1:], start=3):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 3. 设置冻结
    ws.freeze_panes = freeze

    # 4. 设置列宽和百分比格式
    for col_i, col_name in enumerate(df.columns, start=1):
        # 估算宽度
        sample = df.iloc[:, col_i - 1].astype(str).head(50).tolist()
        max_len = max([len(str(col_name))] + [len(x) for x in sample]) if sample else len(str(col_name))
        ws.column_dimensions[get_column_letter(col_i)].width = min(max(10, max_len * 1.2), 60)

        # 百分比格式
        if percent_cols and col_name in percent_cols:
            col_letter = get_column_letter(col_i)
            # Apply to data rows only
            for r in range(3, ws.max_row + 1):
                ws.cell(row=r, column=col_i).number_format = "0.00%"


def write_final_excel(
        out_path: str,
        chunks: List[Chunk],
        records_df: pd.DataFrame,
        unparsed_df: pd.DataFrame,
        tables: Dict[str, pd.DataFrame],
        final_ai: Optional[Dict[str, Any]] = None
):
    print("正在写入 Excel (增强版)...")

    # 复制并清洗数据，防止 JSON 对象报错
    export_records = records_df.copy()
    export_unparsed = unparsed_df.copy()

    def serialize_complex_columns(df: pd.DataFrame):
        if df.empty: return df
        for col in df.columns:
            # 简单判断是否包含 list 或 dict
            if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
                df[col] = df[col].apply(
                    lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, (list, dict)) else x)
        return df

    export_records = serialize_complex_columns(export_records)
    export_unparsed = serialize_complex_columns(export_unparsed)

    wb = Workbook()
    wb.remove(wb.active)

    # --- 0. README ---
    ws = wb.create_sheet("README")
    lines = [
        "统计分析报告：系统功能语言学及物性分析",
        "------------------------------------------------",
        f"数据源：额尔古纳河右岸.pdf (正文)",
        f"统计范围：包含主句及嵌入小句 (Embedded Clauses)",
        "------------------------------------------------",
        "Sheet 说明：",
        "1. Summary_ProcessTypes: 全书及物性过程类型总体分布 (含嵌入小句)",
        "2. Summary_NatureSubcategory: 大自然主动参与者的内部构成 (动物/植物/环境)",
        "3. Summary_NatureVsHuman: 大自然 vs 人类 vs 人造物 的能动性对比",
        "4. NatureActive_RoleDist: 大自然作为不同语义角色 (Actor, Sayer等) 的分布",
        "5. NatureActive_ProcessDist: 大自然作为主动方参与的不同过程类型分布",
        "6. TopN_Animals_Detailed: 各类动物作为主动参与者的详细统计",
        "7. Annotations_All: 所有标注明细数据",
        "------------------------------------------------",
        "统计规则说明：",
        "- 动物部位 (如鹿角) 归并入 动物 (Animal)",
        "- 存在过程 (Existential) 不计入主动参与者 (Active Counted = 0)",
        "- 嵌入小句 (Embedded) 已计入过程类型总量统计",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100
    ws["A1"].font = Font(bold=True, size=14)

    # --- 1. 明细数据 ---
    # 分页写入防止超出限制
    chunk_df = pd.DataFrame([asdict(c) for c in chunks if hasattr(c, 'chunk_id')])  # 简单转换
    if "text_with_markers" in chunk_df.columns:
        chunk_df = chunk_df.drop(columns=["text_with_markers"])

    add_sheet_from_df(wb, "Index_Chunks", chunk_df, description="文本分块索引信息")

    if len(export_records) <= EXCEL_MAX_ROWS:
        add_sheet_from_df(wb, "Annotations_All", export_records,
                          description="[核心数据] 逐句标注明细，包含主句与嵌入小句")
    else:
        # Split logic omitted for brevity, same as original
        add_sheet_from_df(wb, "Annotations_All", export_records.head(EXCEL_MAX_ROWS),
                          description="[核心数据] 逐句标注明细 (Part 1)")

    if not export_unparsed.empty:
        add_sheet_from_df(wb, "Unparsed", export_unparsed, description="未成功解析的片段 (不计入统计)")

    # --- 2. 统计表 (带说明) ---

    # 定义每个表的中文说明映射
    table_descriptions = {
        "Summary_ProcessTypes": "【全书概览】所有及物性过程类型的数量与比例（分母=全书识别出的过程总数，含嵌入过程）。反映小说整体的叙事动态性。",
        "Summary_NatureSubcategory": "【大自然内部结构】在大自然作为主动参与者的事件中，动物、植物、自然环境各自的占比。",
        "Summary_NatureVsHuman": "【能动性对比】全书中“人类”、“大自然”、“人造物”作为主动参与者的数量对比。",
        "NatureActive_RoleDist": "【自然的角色】大自然在事件中充当的语义角色分布。",
        "NatureActive_ProcessDist": "【自然的行动】大自然作为主动方时，其参与的过程类型分布。",
        "Top10_SpeciesActive": "【活跃物种TOP10】全书中作为主动参与者出现频率最高的具体物种。",
        "TopN_Animals_Detailed": "【动物群像】各类动物作为主动参与者的详细统计。",
        # ========== 新增表格说明 ==========
        "Syntax_Ba_Bei": "【特种句式与受事赋权】统计小说中大自然作为物质过程动作者(Actor)时的“把字句”使用比例，以及大自然作为物质过程目标(Goal)时的“被字句”被动赋权比例。",
        "Syntax_Copula_Relational": "【关系与属性建构】统计大自然作为关系过程的载体(Carrier)时，系词('是', '像', '似', '无')的使用分布，这反映了拟人化、比喻或直接描写的叙事偏好。",
        "Circumstances_Dist": "【时空与方式环境】全文环境成分（时间、地点、方式）的数量分布对比，以及大自然占据主导能动性(Active)时的环境成分偏好。"
    }

    table_descriptions.update({
        "Spatial_Types_Dist": "【地点类别与子类】所有的地点环境成分主类型与子类型分布及比例。",
        "Spatial_Nature_Keywords": "【自然地貌词频】自然类地点中，包含‘河,山,树,地,林,水’等字眼的数量与比例。",
        "Spatial_Nature_Top10": "【十大自然地点】出现频次最高的前10个自然类地点（已进行代词消解）。",
        "Spatial_Markers_Dist": "【方位词与介词】地点环境成分中，各类特定方位词与介词的使用频次。",
        "Spatial_Tree_Details": "【树木意象细分】包含‘树’或‘林’的自然地点中，树种的分布及树木相对位置的分布。",
        "Spatial_Nature_OverallSyntax": "【整体句法结构】自然类地点成分的宏观语法结构（如‘介词+处所词+方位词’）。",
        "Spatial_All_NounSyntax": "【处所词微观句法】所有处所词内部的组词语法结构（如‘名+的+名’）。",
        "Spatial_CoOccur_Participants": "【生态互动网络】在出现自然类地点的小句中，主动参与者是人类还是大自然的比例对比。"
    })

    # 定义百分比列，自动进行 Excel 格式化
    percent_map = {
        "Summary_ProcessTypes": ["pct"],
        "Summary_NatureSubcategory": ["pct_of_nature"],
        "Summary_NatureVsHuman": ["pct_of_active"],
        "NatureActive_RoleDist": ["pct_of_nature_active"],
        "NatureActive_ProcessDist": ["pct_of_nature_active"],
        "Top10_SpeciesActive": ["pct_of_flora_fauna_active"],
        "TopN_Animals_Detailed": ["pct_of_all_animals"],
        # ========== 新增百分比列 ==========
        "Syntax_Ba_Bei": ["Ratio"],
        "Syntax_Copula_Relational": ["Ratio"],
        "Circumstances_Dist": ["Total_Ratio", "Nature_Active_Ratio"]
    }

    percent_map.update({
        "Spatial_Types_Dist": ["Ratio"],
        "Spatial_Nature_Keywords": ["Ratio"],
        "Spatial_Nature_Top10": ["Ratio"],
        "Spatial_Markers_Dist": ["Ratio"],
        "Spatial_Tree_Details": ["Species_Ratio", "Position_Ratio"],  # 针对合并的并排表
        "Spatial_Nature_OverallSyntax": ["Ratio"],
        "Spatial_All_NounSyntax": ["Ratio"],
        "Spatial_CoOccur_Participants": ["Ratio"]
    })

    for name, df in tables.items():
        desc = table_descriptions.get(name, "统计数据表")
        pct_cols = percent_map.get(name, [])
        add_sheet_from_df(wb, name, df, percent_cols=pct_cols, description=desc)

    # --- 3. AI Synthesis ---
    if final_ai:
        ws2 = wb.create_sheet("AI_Synthesis_Report")
        ws2.column_dimensions["A"].width = 120
        ws2["A1"] = "AI 定性分析报告 (基于统计数据)"
        ws2["A1"].font = Font(bold=True, size=14)

        row = 3
        ws2[f"A{row}"] = "【叙事综述】"
        ws2[f"A{row}"].font = Font(bold=True)
        ws2[f"A{row + 1}"] = final_ai.get("narrative_summary_cn", "")

        row += 4
        ws2[f"A{row}"] = "【一致性检查与建议】"
        ws2[f"A{row}"].font = Font(bold=True)
        for item in final_ai.get("consistency_flags", []):
            row += 1
            ws2[f"A{row}"] = f"• 问题点: {item.get('issue', '')} (建议核查: {item.get('how_to_audit', '')})"

    wb.save(out_path)


# -----------------------------
# MAIN
# -----------------------------
def main():
    if not API_KEY:
        raise ValueError("API_KEY 为空。请在脚本顶部填入你的 key。")

    ensure_dir(OUT_DIR)
    cache_dir = os.path.join(OUT_DIR, "cache_chunks")
    ensure_dir(cache_dir)

    print("[1] Extract PDF text...")
    full_text = extract_pdf_text(PDF_PATH)

    print("[2] Slice narrative scope (skip author bio)...")
    scope_compact = slice_narrative_scope(full_text, START_SENT, END_SENT)

    print("[3] Sentence split...")
    sentences = sentence_split(scope_compact)
    # Safety: ensure first/last anchor presence
    if not sentences or (START_SENT.replace(" ", "") not in sentences[0].replace(" ", "")):
        print("WARN: 第一条句子未包含起点锚点，可能是PDF抽取差异，但仍继续。")
    if END_SENT.replace(" ", "") not in sentences[-1].replace(" ", ""):
        print("WARN: 最后一条句子未包含终点锚点，可能是PDF抽取差异，但仍继续。")

    print(f"Total sentences: {len(sentences)}")

    print("[4] Build chunks (paragraph-preferred, sentence-safe)...")
    chunks = build_chunks(sentences)
    print(f"Total chunks: {len(chunks)}")
    chunk_map = {c.chunk_id: c for c in chunks}

    print("[5] Call AI per chunk (with cache/resume)...")
    # client = make_client()
    check_ai_connectivity_or_exit()

    payloads = []
    failed_chunks = []
    for ch in chunks:
        try:
            payload = call_ai_for_chunk(ch, cache_dir)
            payloads.append(payload)
            print(
                f"  chunk {ch.chunk_id}: ok, records={len(payload.get('records', []))}, unparsed={len(payload.get('unparsed', []))}")
        except Exception as e:
            failed_chunks.append({"chunk_id": ch.chunk_id, "error": str(e)})
            print(f"  chunk {ch.chunk_id}: FAILED: {e}")

    # Save failures
    if failed_chunks:
        fail_path = os.path.join(OUT_DIR, "failed_chunks.json")
        with open(fail_path, "w", encoding="utf-8") as f:
            json.dump(failed_chunks, f, ensure_ascii=False, indent=2)
        print(f"WARN: some chunks failed. See: {fail_path}")

    if not payloads:
        raise RuntimeError("No successful chunk payloads; abort.")

    print("[6] Flatten records + compute tables...")
    records_df, unparsed_df = flatten_records(payloads, chunk_map)

    # Minimal normalization: ensure required columns exist
    for col in [
        "sent_id", "clause_id", "clause_text", "process_type", "process_subtype", "process_lexeme",
        "active_participant", "participant_category", "nature_subcategory", "canonical_flora_fauna",
        "role", "circumstances", "embedded_processes", "active_counted", "notes", "chunk_id"
    ]:
        if col not in records_df.columns:
            records_df[col] = None

    # Drop unparsed from stats by design (they are in unparsed_df)
    tables = compute_required_tables(records_df)

    print("[7] Final AI synthesis (optional but requested)...")
    final_ai = None
    # try:
    #     final_ai = call_ai_for_final_synthesis(client, tables, OUT_DIR)
    #     print("  final synthesis: ok")
    # except Exception as e:
    #     print(f"WARN: final synthesis failed: {e}")

    print("[8] Write Excel...")
    out_xlsx = os.path.join(OUT_DIR, "Ergun_FullNovel_SFL_Transitivity_v7.xlsx")
    write_final_excel(out_xlsx, chunks, records_df, unparsed_df, tables, final_ai=final_ai)

    print("DONE.")
    print(f"Excel saved to: {out_xlsx}")


if __name__ == "__main__":
    main()
